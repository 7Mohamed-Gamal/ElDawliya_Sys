"""
خدمة التصدير المتقدمة
"""

import os
import json
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
from datetime import datetime, timedelta
from django.conf import settings
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.utils import timezone
from io import BytesIO
import logging

# استيراد مكتبات PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# استيراد مكتبات Excel المتقدمة
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportService:
    """خدمة التصدير المتقدمة"""
    
    def __init__(self):
        self.supported_formats = ['pdf', 'excel', 'csv', 'html', 'json']
        self.export_generators = {
            'pdf': self._export_to_pdf,
            'excel': self._export_to_excel,
            'csv': self._export_to_csv,
            'html': self._export_to_html,
            'json': self._export_to_json,
        }
    
    def export_data(self, data, format_type, template_name, metadata=None):
        """تصدير البيانات بالصيغة المحددة"""
        try:
            if format_type not in self.supported_formats:
                raise ValueError(f"صيغة غير مدعومة: {format_type}")
            
            generator = self.export_generators.get(format_type)
            if not generator:
                raise ValueError(f"مولد غير متاح للصيغة: {format_type}")
            
            return generator(data, template_name, metadata or {})
        except Exception as e:
            logger.error(f"خطأ في تصدير البيانات: {e}")
            raise
    
    def _export_to_pdf(self, data, template_name, metadata):
        """تصدير إلى PDF مع تصميم احترافي"""
        try:
            if not REPORTLAB_AVAILABLE:
                raise ImportError("مكتبة ReportLab غير متاحة")
            
            buffer = BytesIO()
            
            # إعداد الوثيقة
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # تسجيل خط عربي إذا كان متاحاً
            try:
                font_path = os.path.join(settings.STATIC_ROOT or settings.STATICFILES_DIRS[0], 'fonts', 'Cairo-Regular.ttf')
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('Cairo', font_path))
                    arabic_font = 'Cairo'
                else:
                    arabic_font = 'Helvetica'
            except:
                arabic_font = 'Helvetica'
            
            # الأنماط
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=arabic_font,
                fontSize=18,
                spaceAfter=30,
                alignment=1,  # وسط
                textColor=colors.HexColor('#007bff')
            )
            
            header_style = ParagraphStyle(
                'CustomHeader',
                parent=styles['Heading2'],
                fontName=arabic_font,
                fontSize=14,
                spaceAfter=12,
                textColor=colors.HexColor('#333333')
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=arabic_font,
                fontSize=10,
                spaceAfter=6
            )
            
            # محتوى الوثيقة
            story = []
            
            # العنوان
            title = metadata.get('title', template_name)
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 12))
            
            # معلومات التقرير
            info_text = f"تاريخ الإنتاج: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
            if metadata.get('date_range'):
                info_text += f" | الفترة: {metadata['date_range']}"
            if metadata.get('record_count'):
                info_text += f" | عدد السجلات: {metadata['record_count']}"
            
            story.append(Paragraph(info_text, normal_style))
            story.append(Spacer(1, 20))
            
            # الجدول
            if isinstance(data, list) and data:
                # تحويل البيانات إلى جدول
                if isinstance(data[0], dict):
                    # استخراج العناوين
                    headers = list(data[0].keys())
                    table_data = [headers]
                    
                    # إضافة البيانات
                    for row in data:
                        table_data.append([str(row.get(col, '')) for col in headers])
                    
                    # إنشاء الجدول
                    table = Table(table_data)
                    
                    # تنسيق الجدول
                    table.setStyle(TableStyle([
                        # تنسيق الرأس
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), arabic_font),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        
                        # تنسيق البيانات
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                        ('FONTNAME', (0, 1), (-1, -1), arabic_font),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        
                        # الحدود
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        
                        # تلوين الصفوف بالتناوب
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]))
                    
                    story.append(table)
            
            # إضافة ملخص إذا كان متاحاً
            if metadata.get('summary'):
                story.append(Spacer(1, 20))
                story.append(Paragraph("ملخص التقرير", header_style))
                for key, value in metadata['summary'].items():
                    story.append(Paragraph(f"{key}: {value}", normal_style))
            
            # تذييل
            story.append(Spacer(1, 30))
            footer_text = "تم إنتاج هذا التقرير بواسطة نظام الموارد البشرية - الدولية"
            story.append(Paragraph(footer_text, normal_style))
            
            # بناء الوثيقة
            doc.build(story)
            
            buffer.seek(0)
            return buffer.getvalue(), 'application/pdf'
            
        except Exception as e:
            logger.error(f"خطأ في تصدير PDF: {e}")
            raise
    
    def _export_to_excel(self, data, template_name, metadata):
        """تصدير إلى Excel مع تنسيق متقدم"""
        try:
            if not OPENPYXL_AVAILABLE:
                # استخدام pandas كبديل
                return self._export_to_excel_pandas(data, template_name, metadata)
            
            # إنشاء مصنف جديد
            wb = Workbook()
            ws = wb.active
            ws.title = template_name[:31]  # حد أقصى 31 حرف لاسم الورقة
            
            # تنسيق العنوان
            title_font = Font(name='Cairo', size=16, bold=True, color='FFFFFF')
            title_fill = PatternFill(start_color='007BFF', end_color='007BFF', fill_type='solid')
            title_alignment = Alignment(horizontal='center', vertical='center')
            
            # تنسيق الرأس
            header_font = Font(name='Cairo', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='0056B3', end_color='0056B3', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # تنسيق البيانات
            data_font = Font(name='Cairo', size=10)
            data_alignment = Alignment(horizontal='center', vertical='center')
            
            # الحدود
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # إضافة العنوان
            title = metadata.get('title', template_name)
            ws.merge_cells('A1:Z1')
            ws['A1'] = title
            ws['A1'].font = title_font
            ws['A1'].fill = title_fill
            ws['A1'].alignment = title_alignment
            ws.row_dimensions[1].height = 30
            
            # إضافة معلومات التقرير
            info_row = 2
            ws.merge_cells(f'A{info_row}:Z{info_row}')
            info_text = f"تاريخ الإنتاج: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
            if metadata.get('date_range'):
                info_text += f" | الفترة: {metadata['date_range']}"
            ws[f'A{info_row}'] = info_text
            ws[f'A{info_row}'].font = Font(name='Cairo', size=10, italic=True)
            ws[f'A{info_row}'].alignment = Alignment(horizontal='center')
            
            # إضافة البيانات
            if isinstance(data, list) and data:
                start_row = 4
                
                if isinstance(data[0], dict):
                    # إضافة رؤوس الأعمدة
                    headers = list(data[0].keys())
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=start_row, column=col_num, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # إضافة البيانات
                    for row_num, row_data in enumerate(data, start_row + 1):
                        for col_num, header in enumerate(headers, 1):
                            cell = ws.cell(row=row_num, column=col_num, value=str(row_data.get(header, '')))
                            cell.font = data_font
                            cell.alignment = data_alignment
                            cell.border = thin_border
                            
                            # تلوين الصفوف بالتناوب
                            if row_num % 2 == 0:
                                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                    
                    # تعديل عرض الأعمدة
                    for col_num in range(1, len(headers) + 1):
                        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = 15
                    
                    # إضافة رسم بياني إذا كانت البيانات مناسبة
                    if len(data) > 1 and len(headers) >= 2:
                        try:
                            chart = BarChart()
                            chart.title = "رسم بياني للبيانات"
                            chart.style = 10
                            chart.x_axis.title = headers[0]
                            chart.y_axis.title = headers[1] if len(headers) > 1 else "القيم"
                            
                            # البيانات للرسم البياني
                            data_range = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(data), max_col=2)
                            categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(data))
                            
                            chart.add_data(data_range, titles_from_data=True)
                            chart.set_categories(categories)
                            
                            # إضافة الرسم البياني
                            ws.add_chart(chart, f"A{start_row + len(data) + 3}")
                        except:
                            pass  # تجاهل أخطاء الرسم البياني
            
            # حفظ في الذاكرة
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
        except Exception as e:
            logger.error(f"خطأ في تصدير Excel: {e}")
            raise
    
    def _export_to_excel_pandas(self, data, template_name, metadata):
        """تصدير إلى Excel باستخدام pandas"""
        try:
            # تحويل البيانات إلى DataFrame
            if isinstance(data, list) and data:
                df = pd.DataFrame(data)
            else:
                df = pd.DataFrame()
            
            buffer = BytesIO()
            
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # كتابة البيانات
                df.to_excel(writer, sheet_name=template_name[:31], index=False)
                
                # الحصول على الورقة للتنسيق
                worksheet = writer.sheets[template_name[:31]]
                
                # تنسيق الرأس
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='007BFF', end_color='007BFF', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
            
            buffer.seek(0)
            return buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
        except Exception as e:
            logger.error(f"خطأ في تصدير Excel بـ pandas: {e}")
            raise
    
    def _export_to_csv(self, data, template_name, metadata):
        """تصدير إلى CSV"""
        try:
            buffer = BytesIO()
            
            if PANDAS_AVAILABLE and isinstance(data, list) and data:
                # استخدام pandas إذا كان متاحاً
                df = pd.DataFrame(data)
                csv_content = df.to_csv(index=False, encoding='utf-8-sig')
                buffer.write(csv_content.encode('utf-8-sig'))
            else:
                # إنشاء CSV يدوياً بدون pandas
                import csv
                from io import StringIO
                
                if data and isinstance(data[0], dict):
                    # كتابة CSV يدوياً
                    csv_buffer = StringIO()
                    headers = list(data[0].keys())
                    
                    writer = csv.DictWriter(csv_buffer, fieldnames=headers)
                    writer.writeheader()
                    writer.writerows(data)
                    
                    csv_content = csv_buffer.getvalue()
                    buffer.write(csv_content.encode('utf-8-sig'))
                else:
                    buffer.write("لا توجد بيانات".encode('utf-8-sig'))
            
            buffer.seek(0)
            return buffer.getvalue(), 'text/csv'
            
        except Exception as e:
            logger.error(f"خطأ في تصدير CSV: {e}")
            raise
    
    def _export_to_html(self, data, template_name, metadata):
        """تصدير إلى HTML"""
        try:
            context = {
                'title': metadata.get('title', template_name),
                'data': data,
                'metadata': metadata,
                'generated_at': timezone.now(),
            }
            
            html_content = render_to_string('Hr/reports/export_templates/html_export.html', context)
            
            return html_content.encode('utf-8'), 'text/html'
            
        except Exception as e:
            logger.error(f"خطأ في تصدير HTML: {e}")
            raise
    
    def _export_to_json(self, data, template_name, metadata):
        """تصدير إلى JSON"""
        try:
            export_data = {
                'title': metadata.get('title', template_name),
                'generated_at': timezone.now().isoformat(),
                'metadata': metadata,
                'data': data
            }
            
            json_content = json.dumps(export_data, ensure_ascii=False, indent=2)
            
            return json_content.encode('utf-8'), 'application/json'
            
        except Exception as e:
            logger.error(f"خطأ في تصدير JSON: {e}")
            raise
    
    def schedule_export(self, template_id, parameters, format_type, schedule_config, user):
        """جدولة تصدير تقرير"""
        try:
            from ..models_reports import ScheduledReport, ReportTemplate
            
            template = ReportTemplate.objects.get(id=template_id)
            
            # حساب موعد التشغيل التالي
            next_run = self._calculate_next_run(schedule_config)
            
            scheduled_report = ScheduledReport.objects.create(
                template=template,
                name=f"{template.name} - مجدول",
                frequency=schedule_config.get('frequency', 'daily'),
                cron_expression=schedule_config.get('cron_expression'),
                next_run=next_run,
                parameters=parameters,
                output_format=format_type,
                email_recipients=schedule_config.get('email_recipients', []),
                email_subject=schedule_config.get('email_subject', f"تقرير {template.name}"),
                email_body=schedule_config.get('email_body', ''),
                created_by=user
            )
            
            return scheduled_report
            
        except Exception as e:
            logger.error(f"خطأ في جدولة التصدير: {e}")
            raise
    
    def _calculate_next_run(self, schedule_config):
        """حساب موعد التشغيل التالي"""
        try:
            frequency = schedule_config.get('frequency', 'daily')
            now = timezone.now()
            
            if frequency == 'daily':
                return now + timedelta(days=1)
            elif frequency == 'weekly':
                return now + timedelta(weeks=1)
            elif frequency == 'monthly':
                return now + timedelta(days=30)
            elif frequency == 'quarterly':
                return now + timedelta(days=90)
            elif frequency == 'yearly':
                return now + timedelta(days=365)
            else:
                return now + timedelta(days=1)
                
        except Exception as e:
            logger.error(f"خطأ في حساب موعد التشغيل: {e}")
            return timezone.now() + timedelta(days=1)
    
    def send_report_email(self, report_instance, recipients, subject=None, body=None):
        """إرسال التقرير عبر البريد الإلكتروني"""
        try:
            if not recipients:
                raise ValueError("لا توجد عناوين بريد إلكتروني")
            
            # إعداد الرسالة
            subject = subject or f"تقرير {report_instance.template.name}"
            body = body or f"مرفق تقرير {report_instance.template.name} المطلوب."
            
            email = EmailMessage(
                subject=subject,
                body=body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=recipients
            )
            
            # إرفاق الملف إذا كان متاحاً
            if report_instance.file_path and default_storage.exists(report_instance.file_path):
                file_content = default_storage.open(report_instance.file_path).read()
                filename = f"{report_instance.template.name}_{report_instance.created_at.strftime('%Y%m%d')}"
                
                # تحديد امتداد الملف
                extensions = {
                    'pdf': '.pdf',
                    'excel': '.xlsx',
                    'csv': '.csv',
                    'html': '.html',
                    'json': '.json',
                }
                filename += extensions.get(report_instance.output_format, '.txt')
                
                email.attach(filename, file_content)
            
            # إرسال الرسالة
            email.send()
            
            return True
            
        except Exception as e:
            logger.error(f"خطأ في إرسال البريد الإلكتروني: {e}")
            raise
    
    def create_export_template(self, template_type, template_config):
        """إنشاء قالب تصدير مخصص"""
        try:
            # TODO: تنفيذ إنشاء قوالب التصدير المخصصة
            pass
        except Exception as e:
            logger.error(f"خطأ في إنشاء قالب التصدير: {e}")
            raise


# إنشاء مثيل الخدمة
export_service = ExportService()