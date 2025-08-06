"""
خدمة الموظفين الشاملة المحسنة - Employee Service Enhanced
يوفر جميع العمليات المتعلقة بإدارة الموظفين مع دعم النماذج المحسنة والموسعة
"""

from django.db import transaction, models
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.conf import settings
from django.core.cache import cache
from django.db.models import Q, Count, Sum, Avg, F, Case, When, Prefetch
from django.contrib.auth.hashers import check_password, make_password
from decimal import Decimal
from datetime import date, datetime, timedelta
import logging
import json
import hashlib
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import csv
from io import StringIO

logger = logging.getLogger('hr_system')


class EmployeeServiceEnhanced:
    """خدمات إدارة الموظفين الشاملة المحسنة"""
    
    def __init__(self):
        """تهيئة الخدمة"""
        self.cache_timeout = getattr(settings, 'EMPLOYEE_CACHE_TIMEOUT', 3600)
        self.max_bulk_operations = getattr(settings, 'MAX_BULK_OPERATIONS', 1000)
    
    # =============================================================================
    # CORE EMPLOYEE MANAGEMENT METHODS
    # =============================================================================
    
    def create_employee_complete(self, employee_data, education_data=None, 
                               insurance_data=None, vehicle_data=None, 
                               files_data=None, emergency_contacts_data=None,
                               training_data=None, user=None):
        """إنشاء موظف مع جميع البيانات المرتبطة"""
        try:
            with transaction.atomic():
                from ..models.employee.employee_models_enhanced import EmployeeEnhanced
                from ..models.employee.employee_education_models import EmployeeEducationEnhanced
                from ..models.employee.employee_insurance_models import EmployeeInsuranceEnhanced
                from ..models.employee.employee_vehicle_models import EmployeeVehicleEnhanced
                from ..models.employee.employee_file_models import EmployeeFileEnhanced
                from ..models.employee.employee_emergency_contact_models import EmployeeEmergencyContactEnhanced
                from ..models.employee.employee_training_models import EmployeeTrainingEnhanced
                
                # Set created_by if user provided
                if user:
                    employee_data['created_by'] = user
                
                # Create main employee record
                employee = EmployeeEnhanced.objects.create(**employee_data)
                logger.info(f"Created enhanced employee: {employee.employee_id}")
                
                # Create education records
                if education_data:
                    for edu_data in education_data:
                        edu_data['employee'] = employee
                        if user:
                            edu_data['created_by'] = user
                        EmployeeEducationEnhanced.objects.create(**edu_data)
                    logger.info(f"Created {len(education_data)} education records for {employee.employee_id}")
                
                # Create insurance records
                if insurance_data:
                    for ins_data in insurance_data:
                        ins_data['employee'] = employee
                        if user:
                            ins_data['created_by'] = user
                        EmployeeInsuranceEnhanced.objects.create(**ins_data)
                    logger.info(f"Created {len(insurance_data)} insurance records for {employee.employee_id}")
                
                # Create vehicle records
                if vehicle_data:
                    for veh_data in vehicle_data:
                        veh_data['employee'] = employee
                        if user:
                            veh_data['created_by'] = user
                        EmployeeVehicleEnhanced.objects.create(**veh_data)
                    logger.info(f"Created {len(vehicle_data)} vehicle records for {employee.employee_id}")
                
                # Create file records (metadata only, files uploaded separately)
                if files_data:
                    for file_data in files_data:
                        file_data['employee'] = employee
                        if user:
                            file_data['uploaded_by'] = user
                        EmployeeFileEnhanced.objects.create(**file_data)
                    logger.info(f"Created {len(files_data)} file records for {employee.employee_id}")
                
                # Create emergency contact records
                if emergency_contacts_data:
                    for contact_data in emergency_contacts_data:
                        contact_data['employee'] = employee
                        if user:
                            contact_data['created_by'] = user
                        EmployeeEmergencyContactEnhanced.objects.create(**contact_data)
                    logger.info(f"Created {len(emergency_contacts_data)} emergency contact records for {employee.employee_id}")
                
                # Create training records
                if training_data:
                    for training_record in training_data:
                        training_record['employee'] = employee
                        if user:
                            training_record['created_by'] = user
                        EmployeeTrainingEnhanced.objects.create(**training_record)
                    logger.info(f"Created {len(training_data)} training records for {employee.employee_id}")
                
                # Clear cache
                self._clear_employee_cache()
                
                return employee
                
        except Exception as e:
            logger.error(f"Error creating enhanced employee: {e}")
            raise ValidationError(f"خطأ في إنشاء الموظف: {e}")
    
    def update_employee_comprehensive(self, employee_id, update_data, user=None):
        """تحديث شامل لبيانات الموظف"""
        try:
            from ..models.employee.employee_models_enhanced import EmployeeEnhanced
            
            with transaction.atomic():
                employee = EmployeeEnhanced.objects.get(id=employee_id)
                
                # Update basic employee data
                for field, value in update_data.get('employee', {}).items():
                    if hasattr(employee, field):
                        setattr(employee, field, value)
                
                if user:
                    employee.updated_by = user
                
                employee.save()
                
                # Update related data if provided
                if 'education' in update_data:
                    self._update_education_records(employee, update_data['education'], user)
                
                if 'insurance' in update_data:
                    self._update_insurance_records(employee, update_data['insurance'], user)
                
                if 'vehicles' in update_data:
                    self._update_vehicle_records(employee, update_data['vehicles'], user)
                
                if 'files' in update_data:
                    self._update_file_records(employee, update_data['files'], user)
                
                if 'emergency_contacts' in update_data:
                    self._update_emergency_contact_records(employee, update_data['emergency_contacts'], user)
                
                if 'training' in update_data:
                    self._update_training_records(employee, update_data['training'], user)
                
                # Clear cache
                self._clear_employee_cache(employee_id)
                
                logger.info(f"Updated enhanced employee: {employee.employee_id}")
                return employee
                
        except EmployeeEnhanced.DoesNotExist:
            raise ValidationError("الموظف غير موجود")
        except Exception as e:
            logger.error(f"Error updating enhanced employee {employee_id}: {e}")
            raise ValidationError(f"خطأ في تحديث الموظف: {e}")
    
    def calculate_service_years_detailed(self, employee):
        """حساب تفصيلي لسنوات الخدمة مع معلومات إضافية"""
        try:
            today = date.today()
            join_date = employee.join_date
            
            # Calculate years, months, days
            years = today.year - join_date.year
            months = today.month - join_date.month
            days = today.day - join_date.day
            
            if days < 0:
                months -= 1
                # Get days in previous month
                if today.month == 1:
                    prev_month = 12
                    prev_year = today.year - 1
                else:
                    prev_month = today.month - 1
                    prev_year = today.year
                
                from calendar import monthrange
                days_in_prev_month = monthrange(prev_year, prev_month)[1]
                days += days_in_prev_month
            
            if months < 0:
                years -= 1
                months += 12
            
            total_days = (today - join_date).days
            total_years = round(total_days / 365.25, 2)
            
            # Calculate service milestones
            milestones = []
            if years >= 1:
                milestones.append(f"{years} سنة")
            if years >= 5:
                milestones.append("خدمة طويلة (5+ سنوات)")
            if years >= 10:
                milestones.append("خدمة مميزة (10+ سنوات)")
            if years >= 20:
                milestones.append("خدمة استثنائية (20+ سنوات)")
            
            # Calculate next milestone
            next_milestone = None
            if years < 1:
                next_milestone = {"milestone": "سنة واحدة", "days_remaining": 365 - total_days}
            elif years < 5:
                next_milestone = {"milestone": "5 سنوات", "days_remaining": (5 * 365) - total_days}
            elif years < 10:
                next_milestone = {"milestone": "10 سنوات", "days_remaining": (10 * 365) - total_days}
            elif years < 20:
                next_milestone = {"milestone": "20 سنة", "days_remaining": (20 * 365) - total_days}
            
            return {
                'years': years,
                'months': months,
                'days': days,
                'total_days': total_days,
                'total_years': total_years,
                'milestones': milestones,
                'next_milestone': next_milestone,
                'service_category': self._get_service_category(years),
                'is_long_service': years >= 5,
                'is_veteran': years >= 15
            }
            
        except Exception as e:
            logger.error(f"Error calculating service years for {employee.employee_id}: {e}")
            return None
    
    def get_employee_complete_profile(self, employee_id, include_sensitive=False):
        """الحصول على الملف الشخصي الكامل للموظف"""
        try:
            from ..models.employee.employee_models_enhanced import EmployeeEnhanced
            
            # Check cache first
            cache_key = f"employee_profile_{employee_id}_{include_sensitive}"
            cached_profile = cache.get(cache_key)
            if cached_profile:
                return cached_profile
            
            employee = EmployeeEnhanced.objects.select_related(
                'company', 'branch', 'department', 'position', 'direct_manager',
                'created_by', 'updated_by'
            ).prefetch_related(
                'education_records',
                'insurance_records', 
                'vehicle_records',
                'files',
                'emergency_contacts',
                'training_records',
                'subordinates_enhanced'
            ).get(id=employee_id)
            
            # Calculate additional information
            service_details = self.calculate_service_years_detailed(employee)
            
            # Get counts
            education_count = employee.education_records.count()
            insurance_count = employee.insurance_records.count()
            vehicle_count = employee.vehicle_records.count()
            files_count = employee.files.count()
            emergency_contacts_count = employee.emergency_contacts.count()
            training_count = employee.training_records.count()
            subordinates_count = employee.subordinates_enhanced.filter(status='active').count()
            
            # Get expiring documents
            expiring_documents = self._get_expiring_documents(employee)
            
            # Get recent activities
            recent_activities = self._get_recent_activities(employee)
            
            profile = {
                'employee': employee,
                'service_details': service_details,
                'counts': {
                    'education': education_count,
                    'insurance': insurance_count,
                    'vehicles': vehicle_count,
                    'files': files_count,
                    'emergency_contacts': emergency_contacts_count,
                    'training': training_count,
                    'subordinates': subordinates_count,
                },
                'expiring_documents': expiring_documents,
                'recent_activities': recent_activities,
                'alerts': self._get_employee_alerts(employee),
                'performance_summary': self._get_performance_summary(employee),
            }
            
            # Add sensitive data if requested and authorized
            if include_sensitive:
                profile['sensitive_data'] = self._get_sensitive_data(employee)
            
            # Cache for 30 minutes
            cache.set(cache_key, profile, 1800)
            
            return profile
            
        except EmployeeEnhanced.DoesNotExist:
            raise ValidationError("الموظف غير موجود")
        except Exception as e:
            logger.error(f"Error getting enhanced employee profile {employee_id}: {e}")
            raise ValidationError(f"خطأ في جلب ملف الموظف: {e}")
    
    def export_employee_comprehensive_data(self, filters=None, format='excel', 
                                         include_sensitive=False, user=None):
        """تصدير شامل لبيانات الموظفين مع خيارات متقدمة"""
        try:
            from ..models.employee.employee_models_enhanced import EmployeeEnhanced
            
            # Check permissions for sensitive data
            if include_sensitive and not self._can_access_sensitive_data(user):
                raise ValidationError("ليس لديك صلاحية للوصول للبيانات الحساسة")
            
            # Get employees with filters
            employees = EmployeeEnhanced.objects.select_related(
                'company', 'branch', 'department', 'position', 'direct_manager'
            ).prefetch_related(
                'education_records', 'insurance_records', 'vehicle_records'
            ).filter(**(filters or {}))
            
            if format == 'excel':
                return self._export_to_excel_enhanced(employees, include_sensitive)
            elif format == 'csv':
                return self._export_to_csv_enhanced(employees, include_sensitive)
            elif format == 'pdf':
                return self._export_to_pdf_enhanced(employees, include_sensitive)
            else:
                raise ValidationError("تنسيق التصدير غير مدعوم")
                
        except Exception as e:
            logger.error(f"Error exporting enhanced employee data: {e}")
            raise ValidationError(f"خطأ في تصدير البيانات: {e}")
    
    # =============================================================================
    # ADVANCED SEARCH AND FILTERING
    # =============================================================================
    
    def search_employees_advanced(self, search_params, user=None):
        """البحث المتقدم في الموظفين مع فلترة متطورة"""
        try:
            from ..models.employee.employee_models_enhanced import EmployeeEnhanced
            
            # Build cache key
            cache_key = f"employee_search_enhanced_{hashlib.md5(json.dumps(search_params, sort_keys=True).encode()).hexdigest()}"
            cached_result = cache.get(cache_key)
            
            if cached_result:
                return cached_result
            
            queryset = EmployeeEnhanced.objects.select_related(
                'company', 'branch', 'department', 'position', 'direct_manager'
            ).prefetch_related('education_records', 'insurance_records', 'vehicle_records')
            
            # Apply security filters based on user permissions
            queryset = self._apply_security_filters(queryset, user)
            
            # Text search across multiple fields
            if search_params.get('search_text'):
                search_text = search_params['search_text']
                queryset = queryset.filter(
                    Q(first_name__icontains=search_text) |
                    Q(middle_name__icontains=search_text) |
                    Q(last_name__icontains=search_text) |
                    Q(first_name_en__icontains=search_text) |
                    Q(last_name_en__icontains=search_text) |
                    Q(employee_id__icontains=search_text) |
                    Q(employee_number__icontains=search_text) |
                    Q(work_email__icontains=search_text)
                )
            
            # Department filter
            if search_params.get('department_ids'):
                queryset = queryset.filter(department_id__in=search_params['department_ids'])
            
            # Branch filter
            if search_params.get('branch_ids'):
                queryset = queryset.filter(branch_id__in=search_params['branch_ids'])
            
            # Position filter
            if search_params.get('position_ids'):
                queryset = queryset.filter(position_id__in=search_params['position_ids'])
            
            # Employment type filter
            if search_params.get('employment_types'):
                queryset = queryset.filter(employment_type__in=search_params['employment_types'])
            
            # Status filter
            if search_params.get('statuses'):
                queryset = queryset.filter(status__in=search_params['statuses'])
            
            # Gender filter
            if search_params.get('genders'):
                queryset = queryset.filter(gender__in=search_params['genders'])
            
            # Marital status filter
            if search_params.get('marital_statuses'):
                queryset = queryset.filter(marital_status__in=search_params['marital_statuses'])
            
            # Nationality filter
            if search_params.get('nationalities'):
                queryset = queryset.filter(nationality__in=search_params['nationalities'])
            
            # Join date range
            if search_params.get('join_date_from'):
                queryset = queryset.filter(join_date__gte=search_params['join_date_from'])
            if search_params.get('join_date_to'):
                queryset = queryset.filter(join_date__lte=search_params['join_date_to'])
            
            # Age range
            if search_params.get('age_from') or search_params.get('age_to'):
                today = date.today()
                if search_params.get('age_from'):
                    birth_date_to = today - timedelta(days=search_params['age_from'] * 365)
                    queryset = queryset.filter(date_of_birth__lte=birth_date_to)
                if search_params.get('age_to'):
                    birth_date_from = today - timedelta(days=search_params['age_to'] * 365)
                    queryset = queryset.filter(date_of_birth__gte=birth_date_from)
            
            # Salary range
            if search_params.get('salary_from'):
                queryset = queryset.filter(base_salary__gte=search_params['salary_from'])
            if search_params.get('salary_to'):
                queryset = queryset.filter(base_salary__lte=search_params['salary_to'])
            
            # Education level filter
            if search_params.get('education_levels'):
                queryset = queryset.filter(
                    education_records__degree_type__in=search_params['education_levels']
                ).distinct()
            
            # Has manager filter
            if search_params.get('has_manager') is not None:
                if search_params['has_manager']:
                    queryset = queryset.filter(direct_manager__isnull=False)
                else:
                    queryset = queryset.filter(direct_manager__isnull=True)
            
            # Contract expiring soon
            if search_params.get('contract_expiring_days'):
                expiry_date = date.today() + timedelta(days=search_params['contract_expiring_days'])
                queryset = queryset.filter(
                    contract_end_date__lte=expiry_date,
                    contract_end_date__gte=date.today()
                )
            
            # Documents expiring soon
            if search_params.get('documents_expiring_days'):
                expiry_date = date.today() + timedelta(days=search_params['documents_expiring_days'])
                queryset = queryset.filter(
                    Q(passport_expiry_date__lte=expiry_date, passport_expiry_date__gte=date.today()) |
                    Q(visa_expiry_date__lte=expiry_date, visa_expiry_date__gte=date.today())
                )
            
            # Ordering
            order_by = search_params.get('order_by', 'employee_id')
            if search_params.get('order_desc'):
                order_by = f'-{order_by}'
            queryset = queryset.order_by(order_by)
            
            # Pagination
            page = search_params.get('page', 1)
            page_size = min(search_params.get('page_size', 50), 200)  # Max 200 per page
            start = (page - 1) * page_size
            end = start + page_size
            
            total_count = queryset.count()
            employees = list(queryset[start:end])
            
            result = {
                'employees': employees,
                'total_count': total_count,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size,
                'has_next': end < total_count,
                'has_previous': page > 1,
                'search_params': search_params
            }
            
            # Cache for 5 minutes
            cache.set(cache_key, result, 300)
            
            return result
            
        except Exception as e:
            logger.error(f"Error in advanced enhanced employee search: {e}")
            raise ValidationError(f"خطأ في البحث المتقدم: {e}")
    
    # =============================================================================
    # STATISTICS AND ANALYTICS
    # =============================================================================
    
    def get_employee_statistics_enhanced(self, filters=None, user=None):
        """إحصائيات شاملة ومتطورة للموظفين"""
        try:
            from ..models.employee.employee_models_enhanced import EmployeeEnhanced
            
            cache_key = f"employee_stats_enhanced_{hashlib.md5(json.dumps(filters or {}, sort_keys=True).encode()).hexdigest()}"
            cached_stats = cache.get(cache_key)
            
            if cached_stats:
                return cached_stats
            
            queryset = EmployeeEnhanced.objects.filter(**(filters or {}))
            queryset = self._apply_security_filters(queryset, user)
            
            # Basic counts
            total_employees = queryset.count()
            active_employees = queryset.filter(status='active').count()
            inactive_employees = queryset.exclude(status='active').count()
            probation_employees = queryset.filter(status='probation').count()
            
            # By department with enhanced metrics
            by_department = queryset.values('department__name').annotate(
                count=Count('id'),
                avg_salary=Avg('base_salary'),
                min_salary=models.Min('base_salary'),
                max_salary=models.Max('base_salary'),
                avg_service_years=Avg(
                    (timezone.now().date() - F('join_date')) / 365.25
                )
            ).order_by('-count')
            
            # By branch with enhanced metrics
            by_branch = queryset.values('branch__name').annotate(
                count=Count('id'),
                avg_salary=Avg('base_salary'),
                active_count=Count('id', filter=Q(status='active'))
            ).order_by('-count')
            
            # By employment type
            by_employment_type = queryset.values('employment_type').annotate(
                count=Count('id'),
                avg_salary=Avg('base_salary')
            ).order_by('-count')
            
            # By gender with additional metrics
            by_gender = queryset.values('gender').annotate(
                count=Count('id'),
                avg_salary=Avg('base_salary'),
                avg_age=Avg(
                    (timezone.now().date() - F('date_of_birth')) / 365.25
                )
            ).order_by('-count')
            
            # By marital status
            by_marital_status = queryset.values('marital_status').annotate(
                count=Count('id')
            ).order_by('-count')
            
            # By nationality
            by_nationality = queryset.values('nationality').annotate(
                count=Count('id')
            ).order_by('-count')[:10]  # Top 10 nationalities
            
            # Age distribution with more granular ranges
            today = date.today()
            age_ranges = [
                ('18-25', 18, 25),
                ('26-30', 26, 30),
                ('31-35', 31, 35),
                ('36-40', 36, 40),
                ('41-45', 41, 45),
                ('46-50', 46, 50),
                ('51-55', 51, 55),
                ('56-60', 56, 60),
                ('60+', 60, 100)
            ]
            
            age_distribution = []
            for range_name, min_age, max_age in age_ranges:
                min_birth_date = today - timedelta(days=max_age * 365)
                max_birth_date = today - timedelta(days=min_age * 365)
                count = queryset.filter(
                    date_of_birth__gte=min_birth_date,
                    date_of_birth__lte=max_birth_date
                ).count()
                age_distribution.append({
                    'range': range_name,
                    'count': count,
                    'percentage': round((count / total_employees * 100), 2) if total_employees > 0 else 0
                })
            
            # Service years distribution
            service_ranges = [
                ('أقل من 6 أشهر', 0, 0.5),
                ('6 أشهر - سنة', 0.5, 1),
                ('1-2 سنة', 1, 2),
                ('3-5 سنوات', 3, 5),
                ('6-10 سنوات', 6, 10),
                ('11-15 سنة', 11, 15),
                ('16-20 سنة', 16, 20),
                ('أكثر من 20 سنة', 20, 100)
            ]
            
            service_distribution = []
            for range_name, min_years, max_years in service_ranges:
                min_join_date = today - timedelta(days=max_years * 365)
                max_join_date = today - timedelta(days=min_years * 365)
                count = queryset.filter(
                    join_date__gte=min_join_date,
                    join_date__lte=max_join_date
                ).count()
                service_distribution.append({
                    'range': range_name,
                    'count': count,
                    'percentage': round((count / total_employees * 100), 2) if total_employees > 0 else 0
                })
            
            # Salary statistics with percentiles
            salary_stats = queryset.aggregate(
                min_salary=models.Min('base_salary'),
                max_salary=models.Max('base_salary'),
                avg_salary=models.Avg('base_salary'),
                total_salary=models.Sum('base_salary'),
                median_salary=models.Avg('base_salary')  # Simplified median
            )
            
            # Recent activities
            recent_hires = queryset.filter(
                join_date__gte=today - timedelta(days=30)
            ).count()
            
            recent_terminations = queryset.filter(
                termination_date__gte=today - timedelta(days=30)
            ).count()
            
            # Upcoming events
            upcoming_birthdays = queryset.filter(
                date_of_birth__month=today.month,
                date_of_birth__day__gte=today.day,
                date_of_birth__day__lte=today.day + 7
            ).count()
            
            upcoming_contract_expiries = queryset.filter(
                contract_end_date__gte=today,
                contract_end_date__lte=today + timedelta(days=30)
            ).count()
            
            upcoming_document_expiries = queryset.filter(
                Q(passport_expiry_date__gte=today, passport_expiry_date__lte=today + timedelta(days=30)) |
                Q(visa_expiry_date__gte=today, visa_expiry_date__lte=today + timedelta(days=30))
            ).count()
            
            # Performance metrics
            high_performers = queryset.filter(performance_rating='excellent').count()
            needs_improvement = queryset.filter(performance_rating='needs_improvement').count()
            
            # Education statistics
            education_stats = {
                'phd': queryset.filter(education_records__degree_type='phd').distinct().count(),
                'master': queryset.filter(education_records__degree_type='master').distinct().count(),
                'bachelor': queryset.filter(education_records__degree_type='bachelor').distinct().count(),
                'diploma': queryset.filter(education_records__degree_type='diploma').distinct().count(),
                'high_school': queryset.filter(education_records__degree_type='high_school').distinct().count(),
            }
            
            # Vehicle statistics
            vehicle_stats = {
                'company_vehicles': queryset.filter(vehicle_records__vehicle_type='company').distinct().count(),
                'personal_vehicles': queryset.filter(vehicle_records__vehicle_type='personal').distinct().count(),
                'vehicle_allowance': queryset.filter(vehicle_records__vehicle_type='allowance').distinct().count(),
            }
            
            # Insurance coverage
            insurance_stats = {
                'social_insurance': queryset.filter(insurance_records__insurance_type='social').distinct().count(),
                'medical_insurance': queryset.filter(insurance_records__insurance_type='medical').distinct().count(),
                'life_insurance': queryset.filter(insurance_records__insurance_type='life').distinct().count(),
            }
            
            # Compile comprehensive statistics
            stats = {
                'overview': {
                    'total_employees': total_employees,
                    'active_employees': active_employees,
                    'inactive_employees': inactive_employees,
                    'probation_employees': probation_employees,
                    'activity_rate': round((active_employees / total_employees * 100), 2) if total_employees > 0 else 0,
                    'recent_hires': recent_hires,
                    'recent_terminations': recent_terminations,
                },
                'demographics': {
                    'by_department': list(by_department),
                    'by_branch': list(by_branch),
                    'by_employment_type': list(by_employment_type),
                    'by_gender': list(by_gender),
                    'by_marital_status': list(by_marital_status),
                    'by_nationality': list(by_nationality),
                    'age_distribution': age_distribution,
                    'service_distribution': service_distribution,
                },
                'compensation': {
                    'salary_stats': salary_stats,
                    'vehicle_stats': vehicle_stats,
                },
                'qualifications': {
                    'education_stats': education_stats,
                    'insurance_stats': insurance_stats,
                },
                'performance': {
                    'high_performers': high_performers,
                    'needs_improvement': needs_improvement,
                    'performance_rate': round((high_performers / total_employees * 100), 2) if total_employees > 0 else 0,
                },
                'upcoming_events': {
                    'birthdays': upcoming_birthdays,
                    'contract_expiries': upcoming_contract_expiries,
                    'document_expiries': upcoming_document_expiries,
                },
                'generated_at': timezone.now(),
                'filters_applied': filters or {},
            }
            
            # Cache for 15 minutes
            cache.set(cache_key, stats, 900)
            
            return stats
            
        except Exception as e:
            logger.error(f"Error generating enhanced employee statistics: {e}")
            raise ValidationError(f"خطأ في إنتاج الإحصائيات: {e}")
    
    # =============================================================================
    # BULK OPERATIONS
    # =============================================================================
    
    def bulk_update_employees(self, employee_updates, user=None):
        """تحديث مجمع للموظفين"""
        try:
            from ..models.employee.employee_models_enhanced import EmployeeEnhanced
            
            if len(employee_updates) > self.max_bulk_operations:
                raise ValidationError(f"عدد العمليات المجمعة يتجاوز الحد الأقصى ({self.max_bulk_operations})")
            
            updated_count = 0
            errors = []
            
            with transaction.atomic():
                for update_data in employee_updates:
                    try:
                        employee_id = update_data.get('employee_id')
                        if not employee_id:
                            errors.append("معرف الموظف مطلوب")
                            continue
                        
                        employee = EmployeeEnhanced.objects.get(id=employee_id)
                        
                        # Update fields
                        for field, value in update_data.get('fields', {}).items():
                            if hasattr(employee, field):
                                setattr(employee, field, value)
                        
                        if user:
                            employee.updated_by = user
                        
                        employee.save()
                        updated_count += 1
                        
                    except EmployeeEnhanced.DoesNotExist:
                        errors.append(f"الموظف {employee_id} غير موجود")
                    except Exception as e:
                        errors.append(f"خطأ في تحديث الموظف {employee_id}: {e}")
            
            # Clear cache
            self._clear_employee_cache()
            
            return {
                'updated_count': updated_count,
                'errors': errors,
                'success': len(errors) == 0
            }
            
        except Exception as e:
            logger.error(f"Error in bulk employee update: {e}")
            raise ValidationError(f"خطأ في التحديث المجمع: {e}")
    
    def bulk_import_employees(self, import_data, user=None, validate_only=False):
        """استيراد مجمع للموظفين من ملف Excel أو CSV"""
        try:
            from ..models.employee.employee_models_enhanced import EmployeeEnhanced
            
            if len(import_data) > self.max_bulk_operations:
                raise ValidationError(f"عدد السجلات يتجاوز الحد الأقصى ({self.max_bulk_operations})")
            
            created_count = 0
            updated_count = 0
            errors = []
            
            # Validation phase
            for row_num, row_data in enumerate(import_data, 1):
                try:
                    # Validate required fields
                    required_fields = ['first_name', 'last_name', 'employee_id', 'company_id']
                    for field in required_fields:
                        if not row_data.get(field):
                            errors.append(f"الصف {row_num}: الحقل {field} مطلوب")
                    
                    # Validate data types and formats
                    if row_data.get('date_of_birth'):
                        try:
                            datetime.strptime(row_data['date_of_birth'], '%Y-%m-%d')
                        except ValueError:
                            errors.append(f"الصف {row_num}: تنسيق تاريخ الميلاد غير صحيح")
                    
                    if row_data.get('base_salary'):
                        try:
                            float(row_data['base_salary'])
                        except ValueError:
                            errors.append(f"الصف {row_num}: قيمة الراتب غير صحيحة")
                    
                except Exception as e:
                    errors.append(f"الصف {row_num}: خطأ في التحقق - {e}")
            
            if validate_only:
                return {
                    'validation_errors': errors,
                    'is_valid': len(errors) == 0,
                    'total_rows': len(import_data)
                }
            
            if errors:
                raise ValidationError(f"أخطاء في التحقق: {errors}")
            
            # Import phase
            with transaction.atomic():
                for row_num, row_data in enumerate(import_data, 1):
                    try:
                        employee_id = row_data.get('employee_id')
                        
                        # Check if employee exists
                        try:
                            employee = EmployeeEnhanced.objects.get(employee_id=employee_id)
                            # Update existing employee
                            for field, value in row_data.items():
                                if hasattr(employee, field) and value is not None:
                                    setattr(employee, field, value)
                            
                            if user:
                                employee.updated_by = user
                            
                            employee.save()
                            updated_count += 1
                            
                        except EmployeeEnhanced.DoesNotExist:
                            # Create new employee
                            if user:
                                row_data['created_by'] = user
                            
                            employee = EmployeeEnhanced.objects.create(**row_data)
                            created_count += 1
                        
                    except Exception as e:
                        errors.append(f"الصف {row_num}: خطأ في الاستيراد - {e}")
            
            # Clear cache
            self._clear_employee_cache()
            
            return {
                'created_count': created_count,
                'updated_count': updated_count,
                'errors': errors,
                'success': len(errors) == 0,
                'total_processed': created_count + updated_count
            }
            
        except Exception as e:
            logger.error(f"Error in bulk employee import: {e}")
            raise ValidationError(f"خطأ في الاستيراد المجمع: {e}")
    
    # =============================================================================
    # HELPER METHODS
    # =============================================================================
    
    def _clear_employee_cache(self, employee_id=None):
        """مسح ذاكرة التخزين المؤقت للموظفين"""
        if employee_id:
            # Clear specific employee cache
            cache.delete_many([
                f"employee_profile_{employee_id}_True",
                f"employee_profile_{employee_id}_False",
            ])
        else:
            # Clear all employee-related cache
            cache.clear()
    
    def _can_access_sensitive_data(self, user):
        """التحقق من صلاحية الوصول للبيانات الحساسة"""
        if not user:
            return False
        
        if user.is_superuser:
            return True
        
        # Check if user has HR permissions
        return user.has_perm('Hr.view_sensitive_employee_data')
    
    def _apply_security_filters(self, queryset, user):
        """تطبيق فلاتر الأمان حسب صلاحيات المستخدم"""
        if not user:
            return queryset.none()
        
        if user.is_superuser:
            return queryset
        
        # Apply department-based filtering if user has limited access
        if hasattr(user, 'employee_profile'):
            user_department = user.employee_profile.department
            if user_department:
                return queryset.filter(department=user_department)
        
        return queryset
    
    def _get_service_category(self, years):
        """تصنيف الخدمة حسب السنوات"""
        if years < 1:
            return "موظف جديد"
        elif years < 3:
            return "موظف مبتدئ"
        elif years < 5:
            return "موظف متوسط الخبرة"
        elif years < 10:
            return "موظف ذو خبرة"
        elif years < 15:
            return "موظف خبير"
        else:
            return "موظف مخضرم"
    
    def _get_expiring_documents(self, employee):
        """الحصول على الوثائق التي ستنتهي صلاحيتها"""
        expiring = []
        today = date.today()
        warning_days = 90
        
        # Check passport
        if employee.passport_expiry_date:
            days_left = (employee.passport_expiry_date - today).days
            if 0 <= days_left <= warning_days:
                expiring.append({
                    'document': 'جواز السفر',
                    'expiry_date': employee.passport_expiry_date,
                    'days_left': days_left,
                    'urgency': 'high' if days_left <= 30 else 'medium'
                })
        
        # Check visa
        if employee.visa_expiry_date:
            days_left = (employee.visa_expiry_date - today).days
            if 0 <= days_left <= warning_days:
                expiring.append({
                    'document': 'التأشيرة',
                    'expiry_date': employee.visa_expiry_date,
                    'days_left': days_left,
                    'urgency': 'high' if days_left <= 30 else 'medium'
                })
        
        # Check contract
        if employee.contract_end_date:
            days_left = (employee.contract_end_date - today).days
            if 0 <= days_left <= warning_days:
                expiring.append({
                    'document': 'العقد',
                    'expiry_date': employee.contract_end_date,
                    'days_left': days_left,
                    'urgency': 'high' if days_left <= 30 else 'medium'
                })
        
        return sorted(expiring, key=lambda x: x['days_left'])
    
    def _get_recent_activities(self, employee):
        """الحصول على الأنشطة الحديثة للموظف"""
        activities = []
        
        # Recent file uploads
        recent_files = employee.files.filter(
            created_at__gte=timezone.now() - timedelta(days=30)
        ).order_by('-created_at')[:5]
        
        for file in recent_files:
            activities.append({
                'type': 'file_upload',
                'description': f"تم رفع ملف: {file.title}",
                'date': file.created_at,
                'icon': '📄'
            })
        
        # Recent training
        recent_training = employee.training_records.filter(
            created_at__gte=timezone.now() - timedelta(days=30)
        ).order_by('-created_at')[:5]
        
        for training in recent_training:
            activities.append({
                'type': 'training',
                'description': f"تدريب: {training.training_name}",
                'date': training.created_at,
                'icon': '📚'
            })
        
        return sorted(activities, key=lambda x: x['date'], reverse=True)[:10]
    
    def _get_employee_alerts(self, employee):
        """الحصول على تنبيهات الموظف"""
        alerts = []
        
        # Document expiry alerts
        expiring_docs = self._get_expiring_documents(employee)
        for doc in expiring_docs:
            if doc['urgency'] == 'high':
                alerts.append({
                    'type': 'document_expiry',
                    'level': 'danger',
                    'message': f"{doc['document']} ينتهي خلال {doc['days_left']} يوم",
                    'action_required': True
                })
        
        # Performance alerts
        if employee.performance_rating == 'needs_improvement':
            alerts.append({
                'type': 'performance',
                'level': 'warning',
                'message': "الأداء يحتاج لتحسين",
                'action_required': True
            })
        
        # Probation period alert
        if employee.is_in_probation:
            alerts.append({
                'type': 'probation',
                'level': 'info',
                'message': "الموظف في فترة التجربة",
                'action_required': False
            })
        
        return alerts
    
    def _get_performance_summary(self, employee):
        """ملخص أداء الموظف"""
        return {
            'current_rating': employee.performance_rating,
            'last_evaluation': employee.last_evaluation_date,
            'next_evaluation': employee.next_evaluation_date,
            'needs_evaluation': employee.next_evaluation_date and employee.next_evaluation_date <= date.today(),
        }
    
    def _get_sensitive_data(self, employee):
        """الحصول على البيانات الحساسة (مشفرة)"""
        # This would decrypt sensitive fields if needed
        return {
            'national_id_hint': f"***{employee.national_id[-4:]}" if employee.national_id else None,
            'mobile_hint': f"***{employee.mobile_phone[-4:]}" if employee.mobile_phone else None,
            'bank_account_hint': f"***{employee.bank_account_number[-4:]}" if employee.bank_account_number else None,
        }month=today.month,
                date_of_birth__day__gte=today.day,
                date_of_birth__day__lte=today.day + 7
            ).count()
            
            # Contract and document expirations
            contracts_expiring = queryset.filter(
                contract_end_date__gte=today,
                contract_end_date__lte=today + timedelta(days=90)
            ).count()
            
            passports_expiring = queryset.filter(
                passport_expiry_date__gte=today,
                passport_expiry_date__lte=today + timedelta(days=90)
            ).count()
            
            visas_expiring = queryset.filter(
                visa_expiry_date__gte=today,
                visa_expiry_date__lte=today + timedelta(days=90)
            ).count()
            
            # Performance metrics
            performance_distribution = queryset.values('performance_rating').annotate(
                count=Count('id')
            ).order_by('-count')
            
            stats = {
                'overview': {
                    'total_employees': total_employees,
                    'active_employees': active_employees,
                    'inactive_employees': inactive_employees,
                    'probation_employees': probation_employees,
                    'activity_rate': round((active_employees / total_employees * 100), 2) if total_employees > 0 else 0
                },
                'demographics': {
                    'by_department': list(by_department),
                    'by_branch': list(by_branch),
                    'by_employment_type': list(by_employment_type),
                    'by_gender': list(by_gender),
                    'by_marital_status': list(by_marital_status),
                    'by_nationality': list(by_nationality),
                    'age_distribution': age_distribution,
                    'service_distribution': service_distribution,
                },
                'financial': {
                    'salary_stats': salary_stats,
                    'total_payroll': salary_stats['total_salary'] or 0,
                    'avg_salary_by_department': list(by_department),
                },
                'recent_activities': {
                    'recent_hires': recent_hires,
                    'recent_terminations': recent_terminations,
                    'net_change': recent_hires - recent_terminations
                },
                'upcoming_events': {
                    'upcoming_birthdays': upcoming_birthdays,
                    'contracts_expiring': contracts_expiring,
                    'passports_expiring': passports_expiring,
                    'visas_expiring': visas_expiring,
                },
                'performance': {
                    'performance_distribution': list(performance_distribution),
                },
                'generated_at': timezone.now(),
                'filters_applied': filters or {}
            }
            
            # Cache for 1 hour
            cache.set(cache_key, stats, 3600)
            
            return stats
            
        except Exception as e:
            logger.error(f"Error getting enhanced employee statistics: {e}")
            raise ValidationError(f"خطأ في جلب الإحصائيات: {e}")
    
    # =============================================================================
    # BULK OPERATIONS
    # =============================================================================
    
    def bulk_update_employees(self, employee_ids, update_data, user=None):
        """تحديث مجمع للموظفين مع تحسينات الأمان والأداء"""
        try:
            from ..models.employee.employee_models_enhanced import EmployeeEnhanced
            
            if len(employee_ids) > self.max_bulk_operations:
                raise ValidationError(f"عدد العمليات المجمعة يتجاوز الحد الأقصى ({self.max_bulk_operations})")
            
            with transaction.atomic():
                employees = EmployeeEnhanced.objects.filter(id__in=employee_ids)
                
                if not employees.exists():
                    raise ValidationError("لم يتم العثور على موظفين للتحديث")
                
                # Check permissions
                if not self._can_bulk_update(user, employees):
                    raise ValidationError("ليس لديك صلاحية لتحديث هؤلاء الموظفين")
                
                updated_count = 0
                errors = []
                
                # Allowed fields for bulk update
                allowed_fields = [
                    'department', 'position', 'direct_manager',
                    'employment_type', 'status', 'base_salary',
                    'work_schedule', 'branch'
                ]
                
                for employee in employees:
                    try:
                        # Validate each update
                        for field, value in update_data.items():
                            if field not in allowed_fields:
                                errors.append(f"الحقل {field} غير مسموح في التحديث المجمع")
                                continue
                            
                            if hasattr(employee, field):
                                setattr(employee, field, value)
                        
                        if user:
                            employee.updated_by = user
                        
                        employee.save()
                        updated_count += 1
                        
                        logger.info(f"Bulk updated enhanced employee {employee.employee_id} by {user.username if user else 'system'}")
                        
                    except Exception as e:
                        errors.append(f"خطأ في تحديث الموظف {employee.employee_id}: {e}")
                
                # Clear cache
                self._clear_employee_cache()
                
                return {
                    'updated_count': updated_count,
                    'total_count': len(employee_ids),
                    'success_rate': round((updated_count / len(employee_ids) * 100), 2),
                    'errors': errors
                }
                
        except Exception as e:
            logger.error(f"Error in bulk update: {e}")
            raise ValidationError(f"خطأ في التحديث المجمع: {e}")
    
    def bulk_import_employees(self, import_data, user=None, validate_only=False):
        """استيراد مجمع للموظفين من ملف Excel أو CSV"""
        try:
            if len(import_data) > self.max_bulk_operations:
                raise ValidationError(f"عدد السجلات يتجاوز الحد الأقصى ({self.max_bulk_operations})")
            
            validation_errors = []
            valid_records = []
            
            # Validate all records first
            for index, record in enumerate(import_data):
                try:
                    # Validate required fields
                    validation_result = self.validate_employee_data(record, is_import=True)
                    if validation_result['is_valid']:
                        valid_records.append(record)
                    else:
                        validation_errors.append({
                            'row': index + 1,
                            'errors': validation_result['errors']
                        })
                except Exception as e:
                    validation_errors.append({
                        'row': index + 1,
                        'errors': [f"خطأ في التحقق: {e}"]
                    })
            
            if validate_only:
                return {
                    'valid_count': len(valid_records),
                    'invalid_count': len(validation_errors),
                    'validation_errors': validation_errors,
                    'can_import': len(validation_errors) == 0
                }
            
            if validation_errors:
                raise ValidationError(f"يوجد {len(validation_errors)} خطأ في البيانات")
            
            # Import valid records
            imported_count = 0
            import_errors = []
            
            with transaction.atomic():
                for record in valid_records:
                    try:
                        employee = self.create_employee_complete(
                            employee_data=record,
                            user=user
                        )
                        imported_count += 1
                        logger.info(f"Imported employee {employee.employee_id}")
                        
                    except Exception as e:
                        import_errors.append(f"خطأ في استيراد السجل: {e}")
            
            return {
                'imported_count': imported_count,
                'total_count': len(import_data),
                'success_rate': round((imported_count / len(import_data) * 100), 2),
                'import_errors': import_errors
            }
            
        except Exception as e:
            logger.error(f"Error in bulk import: {e}")
            raise ValidationError(f"خطأ في الاستيراد المجمع: {e}")
    
    # =============================================================================
    # VALIDATION AND SECURITY
    # =============================================================================
    
    def validate_employee_data(self, employee_data, is_update=False, is_import=False):
        """التحقق الشامل من صحة بيانات الموظف"""
        errors = []
        warnings = []
        
        try:
            from ..models.employee.employee_models_enhanced import EmployeeEnhanced
            
            # Required fields for new employee
            if not is_update:
                required_fields = [
                    'first_name', 'last_name', 'date_of_birth', 'gender',
                    'company', 'join_date'
                ]
                
                for field in required_fields:
                    if not employee_data.get(field):
                        errors.append(f"الحقل {field} مطلوب")
            
            # Validate employee ID uniqueness
            if employee_data.get('employee_id'):
                existing = EmployeeEnhanced.objects.filter(
                    employee_id=employee_data['employee_id']
                )
                if is_update and employee_data.get('id'):
                    existing = existing.exclude(id=employee_data['id'])
                
                if existing.exists():
                    errors.append("رقم الموظف موجود مسبقاً")
            
            # Validate email uniqueness
            if employee_data.get('work_email'):
                existing = EmployeeEnhanced.objects.filter(
                    work_email=employee_data['work_email']
                )
                if is_update and employee_data.get('id'):
                    existing = existing.exclude(id=employee_data['id'])
                
                if existing.exists():
                    errors.append("البريد الإلكتروني موجود مسبقاً")
            
            # Validate national ID uniqueness (if provided)
            if employee_data.get('national_id'):
                # Note: national_id is encrypted, so we need special handling
                pass  # Will be handled by model validation
            
            # Validate dates
            if employee_data.get('date_of_birth'):
                birth_date = employee_data['date_of_birth']
                if isinstance(birth_date, str):
                    try:
                        birth_date = datetime.strptime(birth_date, '%Y-%m-%d').date()
                    except ValueError:
                        errors.append("تنسيق تاريخ الميلاد غير صحيح")
                        birth_date = None
                
                if birth_date:
                    age = (date.today() - birth_date).days / 365.25
                    if age < 16:
                        errors.append("عمر الموظف يجب أن يكون 16 سنة على الأقل")
                    elif age > 80:
                        warnings.append("عمر الموظف يبدو كبيراً جداً")
            
            # Validate join date
            if employee_data.get('join_date'):
                join_date = employee_data['join_date']
                if isinstance(join_date, str):
                    try:
                        join_date = datetime.strptime(join_date, '%Y-%m-%d').date()
                    except ValueError:
                        errors.append("تنسيق تاريخ التوظيف غير صحيح")
                        join_date = None
                
                if join_date and join_date > date.today():
                    errors.append("تاريخ التوظيف لا يمكن أن يكون في المستقبل")
            
            # Validate salary
            if employee_data.get('base_salary'):
                try:
                    salary = Decimal(str(employee_data['base_salary']))
                    if salary < 0:
                        errors.append("الراتب يجب أن يكون أكبر من الصفر")
                    elif salary > 1000000:  # 1 million
                        warnings.append("الراتب يبدو مرتفعاً جداً")
                except (ValueError, TypeError):
                    errors.append("قيمة الراتب غير صحيحة")
            
            # Validate phone numbers
            phone_fields = ['mobile_phone', 'home_phone']
            for field in phone_fields:
                if employee_data.get(field):
                    phone = employee_data[field]
                    if not re.match(r'^\+?[1-9]\d{1,14}$', phone):
                        errors.append(f"تنسيق {field} غير صحيح")
            
            # Validate email format
            if employee_data.get('work_email'):
                email = employee_data['work_email']
                if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                    errors.append("تنسيق البريد الإلكتروني غير صحيح")
            
            return {
                'is_valid': len(errors) == 0,
                'errors': errors,
                'warnings': warnings
            }
            
        except Exception as e:
            logger.error(f"Error validating employee data: {e}")
            return {
                'is_valid': False,
                'errors': [f"خطأ في التحقق من البيانات: {e}"],
                'warnings': []
            }
    
    # =============================================================================
    # HELPER METHODS
    # =============================================================================
    
    def _get_service_category(self, years):
        """تحديد فئة الخدمة بناءً على سنوات الخدمة"""
        if years < 1:
            return "موظف جديد"
        elif years < 3:
            return "موظف مبتدئ"
        elif years < 7:
            return "موظف متوسط الخبرة"
        elif years < 15:
            return "موظف خبير"
        else:
            return "موظف مخضرم"
    
    def _get_expiring_documents(self, employee):
        """الحصول على الوثائق التي تنتهي صلاحيتها قريباً"""
        expiring = []
        today = date.today()
        warning_days = 90
        
        if employee.passport_expiry_date:
            days_remaining = (employee.passport_expiry_date - today).days
            if 0 <= days_remaining <= warning_days:
                expiring.append({
                    'document': 'جواز السفر',
                    'expiry_date': employee.passport_expiry_date,
                    'days_remaining': days_remaining,
                    'urgency': 'high' if days_remaining <= 30 else 'medium'
                })
        
        if employee.visa_expiry_date:
            days_remaining = (employee.visa_expiry_date - today).days
            if 0 <= days_remaining <= warning_days:
                expiring.append({
                    'document': 'التأشيرة',
                    'expiry_date': employee.visa_expiry_date,
                    'days_remaining': days_remaining,
                    'urgency': 'high' if days_remaining <= 30 else 'medium'
                })
        
        if employee.contract_end_date:
            days_remaining = (employee.contract_end_date - today).days
            if 0 <= days_remaining <= warning_days:
                expiring.append({
                    'document': 'العقد',
                    'expiry_date': employee.contract_end_date,
                    'days_remaining': days_remaining,
                    'urgency': 'high' if days_remaining <= 30 else 'medium'
                })
        
        return expiring
    
    def _get_recent_activities(self, employee):
        """الحصول على الأنشطة الحديثة للموظف"""
        activities = []
        
        # This would typically query an audit log or activity table
        # For now, we'll return basic information
        activities.append({
            'type': 'profile_update',
            'description': 'تحديث الملف الشخصي',
            'date': employee.updated_at,
            'user': employee.updated_by.username if employee.updated_by else 'النظام'
        })
        
        return activities
    
    def _get_employee_alerts(self, employee):
        """الحصول على التنبيهات الخاصة بالموظف"""
        alerts = []
        
        # Check for expiring documents
        expiring_docs = self._get_expiring_documents(employee)
        for doc in expiring_docs:
            if doc['urgency'] == 'high':
                alerts.append({
                    'type': 'document_expiry',
                    'level': 'error',
                    'message': f"{doc['document']} ينتهي خلال {doc['days_remaining']} يوم"
                })
        
        # Check probation period
        if employee.is_in_probation:
            alerts.append({
                'type': 'probation',
                'level': 'warning',
                'message': 'الموظف في فترة التجربة'
            })
        
        # Check missing information
        if not employee.work_email:
            alerts.append({
                'type': 'missing_info',
                'level': 'info',
                'message': 'البريد الإلكتروني للعمل غير محدد'
            })
        
        return alerts
    
    def _get_performance_summary(self, employee):
        """الحصول على ملخص الأداء"""
        return {
            'last_evaluation': employee.last_evaluation_date,
            'next_evaluation': employee.next_evaluation_date,
            'current_rating': employee.performance_rating,
            'needs_evaluation': employee.next_evaluation_date and employee.next_evaluation_date <= date.today()
        }
    
    def _get_sensitive_data(self, employee):
        """الحصول على البيانات الحساسة (مع فك التشفير)"""
        # This would require proper decryption logic
        # For now, return placeholder
        return {
            'has_encrypted_data': True,
            'encrypted_fields': employee._sensitive_fields
        }
    
    def _can_access_sensitive_data(self, user):
        """التحقق من صلاحية الوصول للبيانات الحساسة"""
        if not user:
            return False
        
        # Check user permissions
        return user.has_perm('Hr.view_sensitive_employee_data')
    
    def _can_bulk_update(self, user, employees):
        """التحقق من صلاحية التحديث المجمع"""
        if not user:
            return False
        
        # Check user permissions
        return user.has_perm('Hr.bulk_update_employees')
    
    def _apply_security_filters(self, queryset, user):
        """تطبيق فلاتر الأمان بناءً على صلاحيات المستخدم"""
        if not user:
            return queryset.none()
        
        # Apply department/branch restrictions if needed
        # This would depend on your security model
        return queryset
    
    def _clear_employee_cache(self, employee_id=None):
        """مسح الكاش المتعلق بالموظفين"""
        if employee_id:
            # Clear specific employee cache
            cache.delete_many([
                f"employee_profile_{employee_id}_True",
                f"employee_profile_{employee_id}_False"
            ])
        else:
            # Clear general employee caches
            cache.delete_pattern("employee_*")
    
    # =============================================================================
    # UPDATE HELPER METHODS
    # =============================================================================
    
    def _update_education_records(self, employee, education_data, user=None):
        """تحديث سجلات التعليم"""
        from ..models.employee.employee_education_models import EmployeeEducationEnhanced
        
        # Clear existing records if requested
        if education_data.get('clear_existing'):
            employee.education_records.all().delete()
        
        # Add new records
        for edu_data in education_data.get('records', []):
            if 'id' in edu_data:
                # Update existing
                EmployeeEducationEnhanced.objects.filter(
                    id=edu_data['id'], employee=employee
                ).update(**{k: v for k, v in edu_data.items() if k != 'id'})
            else:
                # Create new
                edu_data['employee'] = employee
                if user:
                    edu_data['created_by'] = user
                EmployeeEducationEnhanced.objects.create(**edu_data)
    
    def _update_insurance_records(self, employee, insurance_data, user=None):
        """تحديث سجلات التأمين"""
        from ..models.employee.employee_insurance_models import EmployeeInsuranceEnhanced
        
        for ins_data in insurance_data.get('records', []):
            if 'id' in ins_data:
                EmployeeInsuranceEnhanced.objects.filter(
                    id=ins_data['id'], employee=employee
                ).update(**{k: v for k, v in ins_data.items() if k != 'id'})
            else:
                ins_data['employee'] = employee
                if user:
                    ins_data['created_by'] = user
                EmployeeInsuranceEnhanced.objects.create(**ins_data)
    
    def _update_vehicle_records(self, employee, vehicle_data, user=None):
        """تحديث سجلات السيارات"""
        from ..models.employee.employee_vehicle_models import EmployeeVehicleEnhanced
        
        for veh_data in vehicle_data.get('records', []):
            if 'id' in veh_data:
                EmployeeVehicleEnhanced.objects.filter(
                    id=veh_data['id'], employee=employee
                ).update(**{k: v for k, v in veh_data.items() if k != 'id'})
            else:
                veh_data['employee'] = employee
                if user:
                    veh_data['created_by'] = user
                EmployeeVehicleEnhanced.objects.create(**veh_data)
    
    def _update_file_records(self, employee, files_data, user=None):
        """تحديث سجلات الملفات"""
        from ..models.employee.employee_file_models import EmployeeFileEnhanced
        
        for file_data in files_data.get('records', []):
            if 'id' in file_data:
                EmployeeFileEnhanced.objects.filter(
                    id=file_data['id'], employee=employee
                ).update(**{k: v for k, v in file_data.items() if k != 'id'})
            else:
                file_data['employee'] = employee
                if user:
                    file_data['uploaded_by'] = user
                EmployeeFileEnhanced.objects.create(**file_data)
    
    def _update_emergency_contact_records(self, employee, contacts_data, user=None):
        """تحديث سجلات جهات الاتصال للطوارئ"""
        from ..models.employee.employee_emergency_contact_models import EmployeeEmergencyContactEnhanced
        
        for contact_data in contacts_data.get('records', []):
            if 'id' in contact_data:
                EmployeeEmergencyContactEnhanced.objects.filter(
                    id=contact_data['id'], employee=employee
                ).update(**{k: v for k, v in contact_data.items() if k != 'id'})
            else:
                contact_data['employee'] = employee
                if user:
                    contact_data['created_by'] = user
                EmployeeEmergencyContactEnhanced.objects.create(**contact_data)
    
    def _update_training_records(self, employee, training_data, user=None):
        """تحديث سجلات التدريب"""
        from ..models.employee.employee_training_models import EmployeeTrainingEnhanced
        
        for training_record in training_data.get('records', []):
            if 'id' in training_record:
                EmployeeTrainingEnhanced.objects.filter(
                    id=training_record['id'], employee=employee
                ).update(**{k: v for k, v in training_record.items() if k != 'id'})
            else:
                training_record['employee'] = employee
                if user:
                    training_record['created_by'] = user
                EmployeeTrainingEnhanced.objects.create(**training_record)
    
    # =============================================================================
    # EXPORT METHODS
    # =============================================================================
    
    def _export_to_excel_enhanced(self, employees, include_sensitive=False):
        """تصدير محسن إلى Excel مع تنسيق متقدم"""
        wb = Workbook()
        
        # Main employee sheet
        ws = wb.active
        ws.title = "الموظفين"
        
        # Headers
        headers = [
            'رقم الموظف', 'الاسم الكامل', 'الاسم بالإنجليزية', 'النوع', 'تاريخ الميلاد',
            'العمر', 'الحالة الاجتماعية', 'الجنسية', 'البريد الإلكتروني', 'الهاتف',
            'الشركة', 'الفرع', 'القسم', 'المنصب', 'تاريخ التوظيف', 'سنوات الخدمة',
            'نوع التوظيف', 'الراتب الأساسي', 'الحالة الوظيفية', 'المدير المباشر'
        ]
        
        if include_sensitive:
            headers.extend(['رقم الهوية', 'رقم الجواز', 'رقم التأمين الاجتماعي'])
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row, employee in enumerate(employees, 2):
            data = [
                employee.employee_id,
                employee.full_name,
                employee.full_name_en or '',
                employee.get_gender_display(),
                employee.date_of_birth,
                employee.age,
                employee.get_marital_status_display(),
                employee.nationality or '',
                employee.work_email or '',
                employee.mobile_phone or '',
                employee.company.name if employee.company else '',
                employee.branch.name if employee.branch else '',
                employee.department.name if employee.department else '',
                employee.position.title if employee.position else '',
                employee.join_date,
                employee.years_of_service,
                employee.get_employment_type_display(),
                float(employee.base_salary) if employee.base_salary else 0,
                employee.get_status_display(),
                employee.direct_manager.full_name if employee.direct_manager else ''
            ]
            
            if include_sensitive:
                # Note: These would need proper decryption
                data.extend(['***', '***', '***'])
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                
                # Format dates
                if isinstance(value, date):
                    cell.number_format = 'DD/MM/YYYY'
                # Format numbers
                elif isinstance(value, (int, float)) and col in [18]:  # Salary column
                    cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        summary_ws = wb.create_sheet("الملخص")
        summary_data = [
            ['إجمالي الموظفين', employees.count()],
            ['الموظفين النشطين', employees.filter(status='active').count()],
            ['متوسط العمر', round(sum(emp.age for emp in employees) / employees.count(), 1)],
            ['متوسط سنوات الخدمة', round(sum(emp.years_of_service for emp in employees) / employees.count(), 1)],
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=row, column=2, value=value)
        
        return wb
    
    def _export_to_csv_enhanced(self, employees, include_sensitive=False):
        """تصدير محسن إلى CSV"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            'رقم الموظف', 'الاسم الكامل', 'الاسم بالإنجليزية', 'النوع', 'تاريخ الميلاد',
            'العمر', 'الحالة الاجتماعية', 'الجنسية', 'البريد الإلكتروني', 'الهاتف',
            'الشركة', 'الفرع', 'القسم', 'المنصب', 'تاريخ التوظيف', 'سنوات الخدمة',
            'نوع التوظيف', 'الراتب الأساسي', 'الحالة الوظيفية', 'المدير المباشر'
        ]
        
        if include_sensitive:
            headers.extend(['رقم الهوية', 'رقم الجواز', 'رقم التأمين الاجتماعي'])
        
        writer.writerow(headers)
        
        # Data rows
        for employee in employees:
            data = [
                employee.employee_id,
                employee.full_name,
                employee.full_name_en or '',
                employee.get_gender_display(),
                employee.date_of_birth,
                employee.age,
                employee.get_marital_status_display(),
                employee.nationality or '',
                employee.work_email or '',
                employee.mobile_phone or '',
                employee.company.name if employee.company else '',
                employee.branch.name if employee.branch else '',
                employee.department.name if employee.department else '',
                employee.position.title if employee.position else '',
                employee.join_date,
                employee.years_of_service,
                employee.get_employment_type_display(),
                float(employee.base_salary) if employee.base_salary else 0,
                employee.get_status_display(),
                employee.direct_manager.full_name if employee.direct_manager else ''
            ]
            
            if include_sensitive:
                data.extend(['***', '***', '***'])
            
            writer.writerow(data)
        
        return output.getvalue()
    
    def _export_to_pdf_enhanced(self, employees, include_sensitive=False):
        """تصدير محسن إلى PDF"""
        # This would require a PDF library like ReportLab
        # For now, return a placeholder
        return "PDF export not implemented yet"


# Create a singleton instance
employee_service_enhanced = EmployeeServiceEnhanced()