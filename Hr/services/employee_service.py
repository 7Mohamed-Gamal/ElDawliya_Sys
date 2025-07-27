"""
Employee Service - خدمات إدارة الموظفين الشاملة المحسنة
"""

from django.db import transaction, models
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.conf import settings
from django.core.cache import cache
from django.db.models import Q, Count, Sum, Avg, F, Case, When
from decimal import Decimal
from datetime import date, datetime, timedelta
import logging
import json
import hashlib

logger = logging.getLogger('hr_system')


class EmployeeService:
    """خدمات إدارة الموظفين الشاملة"""
    
    def create_employee_complete(self, employee_data, education_data=None, 
                               insurance_data=None, vehicle_data=None, files_data=None):
        """إنشاء موظف مع جميع البيانات المرتبطة"""
        try:
            with transaction.atomic():
                from ..models_enhanced import Employee, EmployeeEducation, EmployeeInsurance, EmployeeVehicle, EmployeeFile
                
                # Create employee
                employee = Employee.objects.create(**employee_data)
                logger.info(f"Created employee: {employee.employee_number}")
                
                # Create education records
                if education_data:
                    for edu_data in education_data:
                        edu_data['employee'] = employee
                        EmployeeEducation.objects.create(**edu_data)
                    logger.info(f"Created {len(education_data)} education records for {employee.employee_number}")
                
                # Create insurance records
                if insurance_data:
                    for ins_data in insurance_data:
                        ins_data['employee'] = employee
                        EmployeeInsurance.objects.create(**ins_data)
                    logger.info(f"Created {len(insurance_data)} insurance records for {employee.employee_number}")
                
                # Create vehicle records
                if vehicle_data:
                    for veh_data in vehicle_data:
                        veh_data['employee'] = employee
                        EmployeeVehicle.objects.create(**veh_data)
                    logger.info(f"Created {len(vehicle_data)} vehicle records for {employee.employee_number}")
                
                # Create file records (metadata only, files uploaded separately)
                if files_data:
                    for file_data in files_data:
                        file_data['employee'] = employee
                        EmployeeFile.objects.create(**file_data)
                    logger.info(f"Created {len(files_data)} file records for {employee.employee_number}")
                
                # Update job position current count
                employee.job_position.update_current_positions()
                
                return employee
                
        except Exception as e:
            logger.error(f"Error creating employee: {e}")
            raise ValidationError(f"خطأ في إنشاء الموظف: {e}")
    
    def update_employee_comprehensive(self, employee_id, update_data):
        """تحديث شامل لبيانات الموظف"""
        try:
            from ..models_enhanced import Employee
            
            employee = Employee.objects.get(id=employee_id)
            
            # Update basic employee data
            for field, value in update_data.get('employee', {}).items():
                setattr(employee, field, value)
            
            employee.save()
            
            # Update related data if provided
            if 'education' in update_data:
                self._update_education_records(employee, update_data['education'])
            
            if 'insurance' in update_data:
                self._update_insurance_records(employee, update_data['insurance'])
            
            if 'vehicles' in update_data:
                self._update_vehicle_records(employee, update_data['vehicles'])
            
            logger.info(f"Updated employee: {employee.employee_number}")
            return employee
            
        except Employee.DoesNotExist:
            raise ValidationError("الموظف غير موجود")
        except Exception as e:
            logger.error(f"Error updating employee {employee_id}: {e}")
            raise ValidationError(f"خطأ في تحديث الموظف: {e}")
    
    def calculate_service_years_detailed(self, employee):
        """حساب تفصيلي لسنوات الخدمة"""
        try:
            from datetime import date
            
            today = date.today()
            hire_date = employee.hire_date
            
            # Calculate years, months, days
            years = today.year - hire_date.year
            months = today.month - hire_date.month
            days = today.day - hire_date.day
            
            if days < 0:
                months -= 1
                # Get days in previous month
                if today.month == 1:
                    prev_month = 12
                    prev_year = today.year - 1
                else:
                    prev_month = today.month - 1
                    prev_year = today.year
                
                from calendar import monthrange
                days_in_prev_month = monthrange(prev_year, prev_month)[1]
                days += days_in_prev_month
            
            if months < 0:
                years -= 1
                months += 12
            
            return {
                'years': years,
                'months': months,
                'days': days,
                'total_days': (today - hire_date).days,
                'total_years': round((today - hire_date).days / 365.25, 2)
            }
            
        except Exception as e:
            logger.error(f"Error calculating service years for {employee.employee_number}: {e}")
            return None
    
    def get_employee_complete_profile(self, employee_id):
        """الحصول على الملف الشخصي الكامل للموظف"""
        try:
            from ..models_enhanced import Employee
            
            employee = Employee.objects.select_related(
                'company', 'branch', 'department', 'job_position', 'direct_manager'
            ).prefetch_related(
                'education_records',
                'insurance_records', 
                'vehicle_records',
                'files',
                'subordinates'
            ).get(id=employee_id)
            
            # Calculate additional information
            service_details = self.calculate_service_years_detailed(employee)
            
            profile = {
                'employee': employee,
                'service_details': service_details,
                'education_count': employee.education_records.count(),
                'insurance_count': employee.insurance_records.count(),
                'vehicle_count': employee.vehicle_records.count(),
                'files_count': employee.files.count(),
                'subordinates_count': employee.subordinates.filter(is_active=True).count(),
            }
            
            return profile
            
        except Employee.DoesNotExist:
            raise ValidationError("الموظف غير موجود")
        except Exception as e:
            logger.error(f"Error getting employee profile {employee_id}: {e}")
            raise ValidationError(f"خطأ في جلب ملف الموظف: {e}")
    
    def export_employee_comprehensive_data(self, filters=None, format='excel'):
        """تصدير شامل لبيانات الموظفين"""
        try:
            from ..models_enhanced import Employee
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Get employees with filters
            employees = Employee.objects.select_related(
                'company', 'branch', 'department', 'job_position'
            ).filter(**(filters or {}))
            
            if format == 'excel':
                return self._export_to_excel(employees)
            elif format == 'csv':
                return self._export_to_csv(employees)
            else:
                raise ValidationError("تنسيق التصدير غير مدعوم")
                
        except Exception as e:
            logger.error(f"Error exporting employee data: {e}")
            raise ValidationError(f"خطأ في تصدير البيانات: {e}")
    
    def _update_education_records(self, employee, education_data):
        """تحديث سجلات التعليم"""
        from ..models_enhanced import EmployeeEducation
        
        # Clear existing records if requested
        if education_data.get('clear_existing'):
            employee.education_records.all().delete()
        
        # Add new records
        for edu_data in education_data.get('records', []):
            if 'id' in edu_data:
                # Update existing
                EmployeeEducation.objects.filter(
                    id=edu_data['id'], employee=employee
                ).update(**{k: v for k, v in edu_data.items() if k != 'id'})
            else:
                # Create new
                edu_data['employee'] = employee
                EmployeeEducation.objects.create(**edu_data)
    
    def _update_insurance_records(self, employee, insurance_data):
        """تحديث سجلات التأمين"""
        from ..models_enhanced import EmployeeInsurance
        
        for ins_data in insurance_data.get('records', []):
            if 'id' in ins_data:
                EmployeeInsurance.objects.filter(
                    id=ins_data['id'], employee=employee
                ).update(**{k: v for k, v in ins_data.items() if k != 'id'})
            else:
                ins_data['employee'] = employee
                EmployeeInsurance.objects.create(**ins_data)
    
    def _update_vehicle_records(self, employee, vehicle_data):
        """تحديث سجلات السيارات"""
        from ..models_enhanced import EmployeeVehicle
        
        for veh_data in vehicle_data.get('records', []):
            if 'id' in veh_data:
                EmployeeVehicle.objects.filter(
                    id=veh_data['id'], employee=employee
                ).update(**{k: v for k, v in veh_data.items() if k != 'id'})
            else:
                veh_data['employee'] = employee
                EmployeeVehicle.objects.create(**veh_data)
    
    def _export_to_excel(self, employees):
        """تصدير إلى Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "قائمة الموظفين"
        
        # Headers
        headers = [
            'رقم الموظف', 'الاسم الكامل', 'البريد الإلكتروني', 'الهاتف',
            'الشركة', 'الفرع', 'القسم', 'المنصب', 'تاريخ التوظيف',
            'نوع التوظيف', 'الراتب الأساسي', 'الحالة'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, employee in enumerate(employees, 2):
            ws.cell(row=row, column=1, value=employee.employee_number)
            ws.cell(row=row, column=2, value=employee.full_name)
            ws.cell(row=row, column=3, value=employee.email)
            ws.cell(row=row, column=4, value=employee.phone_primary)
            ws.cell(row=row, column=5, value=employee.company.name)
            ws.cell(row=row, column=6, value=employee.branch.name)
            ws.cell(row=row, column=7, value=employee.department.name)
            ws.cell(row=row, column=8, value=employee.job_position.title)
            ws.cell(row=row, column=9, value=employee.hire_date)
            ws.cell(row=row, column=10, value=employee.get_employment_type_display())
            ws.cell(row=row, column=11, value=float(employee.basic_salary))
            ws.cell(row=row, column=12, value=employee.get_status_display())
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb
    
    def _export_to_csv(self, employees):
        """تصدير إلى CSV"""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            'رقم الموظف', 'الاسم الكامل', 'البريد الإلكتروني', 'الهاتف',
            'الشركة', 'الفرع', 'القسم', 'المنصب', 'تاريخ التوظيف',
            'نوع التوظيف', 'الراتب الأساسي', 'الحالة'
        ])
        
        # Data rows
        for employee in employees:
            writer.writerow([
                employee.employee_number,
                employee.full_name,
                employee.email,
                employee.phone_primary,
                employee.company.name,
                employee.branch.name,
                employee.department.name,
                employee.job_position.title,
                employee.hire_date,
                employee.get_employment_type_display(),
                employee.basic_salary,
                employee.get_status_display()
            ])
        
        return output.getvalue()
  
  # =============================================================================
    # ADVANCED EMPLOYEE MANAGEMENT METHODS
    # =============================================================================
    
    def search_employees_advanced(self, search_params):
        """البحث المتقدم في الموظفين"""
        try:
            from ..models import Employee
            
            # Build cache key
            cache_key = f"employee_search_{hashlib.md5(json.dumps(search_params, sort_keys=True).encode()).hexdigest()}"
            cached_result = cache.get(cache_key)
            
            if cached_result:
                return cached_result
            
            queryset = Employee.objects.select_related(
                'company', 'branch', 'department', 'job_position', 'direct_manager'
            ).prefetch_related('education_records', 'insurance_records')
            
            # Text search
            if search_params.get('search_text'):
                search_text = search_params['search_text']
                queryset = queryset.filter(
                    Q(first_name__icontains=search_text) |
                    Q(last_name__icontains=search_text) |
                    Q(employee_number__icontains=search_text) |
                    Q(email__icontains=search_text) |
                    Q(phone_primary__icontains=search_text) |
                    Q(national_id__icontains=search_text)
                )
            
            # Department filter
            if search_params.get('department_ids'):
                queryset = queryset.filter(department_id__in=search_params['department_ids'])
            
            # Branch filter
            if search_params.get('branch_ids'):
                queryset = queryset.filter(branch_id__in=search_params['branch_ids'])
            
            # Job position filter
            if search_params.get('job_position_ids'):
                queryset = queryset.filter(job_position_id__in=search_params['job_position_ids'])
            
            # Employment type filter
            if search_params.get('employment_types'):
                queryset = queryset.filter(employment_type__in=search_params['employment_types'])
            
            # Status filter
            if search_params.get('statuses'):
                queryset = queryset.filter(status__in=search_params['statuses'])
            
            # Hire date range
            if search_params.get('hire_date_from'):
                queryset = queryset.filter(hire_date__gte=search_params['hire_date_from'])
            if search_params.get('hire_date_to'):
                queryset = queryset.filter(hire_date__lte=search_params['hire_date_to'])
            
            # Age range
            if search_params.get('age_from') or search_params.get('age_to'):
                today = date.today()
                if search_params.get('age_from'):
                    birth_date_to = today - timedelta(days=search_params['age_from'] * 365)
                    queryset = queryset.filter(birth_date__lte=birth_date_to)
                if search_params.get('age_to'):
                    birth_date_from = today - timedelta(days=search_params['age_to'] * 365)
                    queryset = queryset.filter(birth_date__gte=birth_date_from)
            
            # Salary range
            if search_params.get('salary_from'):
                queryset = queryset.filter(basic_salary__gte=search_params['salary_from'])
            if search_params.get('salary_to'):
                queryset = queryset.filter(basic_salary__lte=search_params['salary_to'])
            
            # Education level filter
            if search_params.get('education_levels'):
                queryset = queryset.filter(
                    education_records__education_type__in=search_params['education_levels']
                ).distinct()
            
            # Has manager filter
            if search_params.get('has_manager') is not None:
                if search_params['has_manager']:
                    queryset = queryset.filter(direct_manager__isnull=False)
                else:
                    queryset = queryset.filter(direct_manager__isnull=True)
            
            # Ordering
            order_by = search_params.get('order_by', 'employee_number')
            if search_params.get('order_desc'):
                order_by = f'-{order_by}'
            queryset = queryset.order_by(order_by)
            
            # Pagination
            page = search_params.get('page', 1)
            page_size = search_params.get('page_size', 50)
            start = (page - 1) * page_size
            end = start + page_size
            
            total_count = queryset.count()
            employees = list(queryset[start:end])
            
            result = {
                'employees': employees,
                'total_count': total_count,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size
            }
            
            # Cache for 5 minutes
            cache.set(cache_key, result, 300)
            
            return result
            
        except Exception as e:
            logger.error(f"Error in advanced employee search: {e}")
            raise ValidationError(f"خطأ في البحث المتقدم: {e}")
    
    def get_employee_statistics(self, filters=None):
        """إحصائيات شاملة للموظفين"""
        try:
            from ..models import Employee
            
            cache_key = f"employee_stats_{hashlib.md5(json.dumps(filters or {}, sort_keys=True).encode()).hexdigest()}"
            cached_stats = cache.get(cache_key)
            
            if cached_stats:
                return cached_stats
            
            queryset = Employee.objects.filter(**(filters or {}))
            
            # Basic counts
            total_employees = queryset.count()
            active_employees = queryset.filter(status='active').count()
            inactive_employees = queryset.filter(status='inactive').count()
            
            # By department
            by_department = queryset.values('department__name').annotate(
                count=Count('id'),
                avg_salary=Avg('basic_salary')
            ).order_by('-count')
            
            # By branch
            by_branch = queryset.values('branch__name').annotate(
                count=Count('id'),
                avg_salary=Avg('basic_salary')
            ).order_by('-count')
            
            # By employment type
            by_employment_type = queryset.values('employment_type').annotate(
                count=Count('id')
            ).order_by('-count')
            
            # By gender
            by_gender = queryset.values('gender').annotate(
                count=Count('id')
            ).order_by('-count')
            
            # Age distribution
            today = date.today()
            age_ranges = [
                ('18-25', 18, 25),
                ('26-35', 26, 35),
                ('36-45', 36, 45),
                ('46-55', 46, 55),
                ('56+', 56, 100)
            ]
            
            age_distribution = []
            for range_name, min_age, max_age in age_ranges:
                min_birth_date = today - timedelta(days=max_age * 365)
                max_birth_date = today - timedelta(days=min_age * 365)
                count = queryset.filter(
                    birth_date__gte=min_birth_date,
                    birth_date__lte=max_birth_date
                ).count()
                age_distribution.append({
                    'range': range_name,
                    'count': count
                })
            
            # Service years distribution
            service_ranges = [
                ('أقل من سنة', 0, 1),
                ('1-3 سنوات', 1, 3),
                ('4-7 سنوات', 4, 7),
                ('8-15 سنة', 8, 15),
                ('أكثر من 15 سنة', 15, 100)
            ]
            
            service_distribution = []
            for range_name, min_years, max_years in service_ranges:
                min_hire_date = today - timedelta(days=max_years * 365)
                max_hire_date = today - timedelta(days=min_years * 365)
                count = queryset.filter(
                    hire_date__gte=min_hire_date,
                    hire_date__lte=max_hire_date
                ).count()
                service_distribution.append({
                    'range': range_name,
                    'count': count
                })
            
            # Salary statistics
            salary_stats = queryset.aggregate(
                min_salary=models.Min('basic_salary'),
                max_salary=models.Max('basic_salary'),
                avg_salary=models.Avg('basic_salary'),
                total_salary=models.Sum('basic_salary')
            )
            
            # Recent hires (last 30 days)
            recent_hires = queryset.filter(
                hire_date__gte=today - timedelta(days=30)
            ).count()
            
            # Upcoming birthdays (next 30 days)
            next_month = today + timedelta(days=30)
            upcoming_birthdays = queryset.filter(
                birth_date__month__gte=today.month,
                birth_date__day__gte=today.day,
                birth_date__month__lte=next_month.month,
                birth_date__day__lte=next_month.day
            ).count()
            
            stats = {
                'total_employees': total_employees,
                'active_employees': active_employees,
                'inactive_employees': inactive_employees,
                'by_department': list(by_department),
                'by_branch': list(by_branch),
                'by_employment_type': list(by_employment_type),
                'by_gender': list(by_gender),
                'age_distribution': age_distribution,
                'service_distribution': service_distribution,
                'salary_stats': salary_stats,
                'recent_hires': recent_hires,
                'upcoming_birthdays': upcoming_birthdays,
                'generated_at': timezone.now()
            }
            
            # Cache for 1 hour
            cache.set(cache_key, stats, 3600)
            
            return stats
            
        except Exception as e:
            logger.error(f"Error getting employee statistics: {e}")
            raise ValidationError(f"خطأ في جلب الإحصائيات: {e}")
    
    def bulk_update_employees(self, employee_ids, update_data, user):
        """تحديث مجمع للموظفين"""
        try:
            from ..models import Employee
            
            with transaction.atomic():
                employees = Employee.objects.filter(id__in=employee_ids)
                
                if not employees.exists():
                    raise ValidationError("لم يتم العثور على موظفين للتحديث")
                
                updated_count = 0
                errors = []
                
                for employee in employees:
                    try:
                        # Update allowed fields only
                        allowed_fields = [
                            'department', 'job_position', 'direct_manager',
                            'employment_type', 'status', 'basic_salary'
                        ]
                        
                        for field, value in update_data.items():
                            if field in allowed_fields:
                                setattr(employee, field, value)
                        
                        employee.save()
                        updated_count += 1
                        
                        logger.info(f"Bulk updated employee {employee.employee_number} by {user.username}")
                        
                    except Exception as e:
                        errors.append(f"خطأ في تحديث الموظف {employee.employee_number}: {e}")
                
                return {
                    'updated_count': updated_count,
                    'total_count': len(employee_ids),
                    'errors': errors
                }
                
        except Exception as e:
            logger.error(f"Error in bulk update: {e}")
            raise ValidationError(f"خطأ في التحديث المجمع: {e}")
    
    def generate_employee_report(self, report_type, filters=None, format='pdf'):
        """إنتاج تقارير شاملة للموظفين"""
        try:
            from ..models import Employee
            
            employees = Employee.objects.select_related(
                'company', 'branch', 'department', 'job_position'
            ).filter(**(filters or {}))
            
            if report_type == 'detailed_list':
                return self._generate_detailed_list_report(employees, format)
            elif report_type == 'summary':
                return self._generate_summary_report(employees, format)
            elif report_type == 'birthday_list':
                return self._generate_birthday_report(employees, format)
            elif report_type == 'service_years':
                return self._generate_service_years_report(employees, format)
            else:
                raise ValidationError("نوع التقرير غير مدعوم")
                
        except Exception as e:
            logger.error(f"Error generating employee report: {e}")
            raise ValidationError(f"خطأ في إنتاج التقرير: {e}")
    
    def validate_employee_data(self, employee_data, is_update=False):
        """التحقق من صحة بيانات الموظف"""
        errors = {}
        
        try:
            from ..models import Employee
            
            # Required fields for new employee
            if not is_update:
                required_fields = [
                    'first_name', 'last_name', 'employee_number',
                    'company', 'branch', 'department', 'job_position',
                    'hire_date', 'basic_salary'
                ]
                
                for field in required_fields:
                    if not employee_data.get(field):
                        errors[field] = f"الحقل {field} مطلوب"
            
            # Validate employee number uniqueness
            if employee_data.get('employee_number'):
                existing = Employee.objects.filter(
                    employee_number=employee_data['employee_number']
                )
                if is_update and employee_data.get('id'):
                    existing = existing.exclude(id=employee_data['id'])
                
                if existing.exists():
                    errors['employee_number'] = "رقم الموظف موجود مسبقاً"
            
            # Validate email uniqueness
            if employee_data.get('email'):
                existing = Employee.objects.filter(email=employee_data['email'])
                if is_update and employee_data.get('id'):
                    existing = existing.exclude(id=employee_data['id'])
                
                if existing.exists():
                    errors['email'] = "البريد الإلكتروني موجود مسبقاً"
            
            # Validate national ID uniqueness
            if employee_data.get('national_id'):
                existing = Employee.objects.filter(national_id=employee_data['national_id'])
                if is_update and employee_data.get('id'):
                    existing = existing.exclude(id=employee_data['id'])
                
                if existing.exists():
                    errors['national_id'] = "رقم الهوية موجود مسبقاً"
            
            # Validate dates
            if employee_data.get('birth_date') and employee_data.get('hire_date'):
                birth_date = employee_data['birth_date']
                hire_date = employee_data['hire_date']
                
                if isinstance(birth_date, str):
                    birth_date = datetime.strptime(birth_date, '%Y-%m-%d').date()
                if isinstance(hire_date, str):
                    hire_date = datetime.strptime(hire_date, '%Y-%m-%d').date()
                
                # Check minimum age (18 years)
                min_age_date = hire_date - timedelta(days=18 * 365)
                if birth_date > min_age_date:
                    errors['birth_date'] = "عمر الموظف يجب أن يكون 18 سنة على الأقل عند التوظيف"
                
                # Check hire date not in future
                if hire_date > date.today():
                    errors['hire_date'] = "تاريخ التوظيف لا يمكن أن يكون في المستقبل"
            
            # Validate salary
            if employee_data.get('basic_salary'):
                try:
                    salary = Decimal(str(employee_data['basic_salary']))
                    if salary <= 0:
                        errors['basic_salary'] = "الراتب الأساسي يجب أن يكون أكبر من صفر"
                except (ValueError, TypeError):
                    errors['basic_salary'] = "الراتب الأساسي يجب أن يكون رقماً صحيحاً"
            
            return {
                'is_valid': len(errors) == 0,
                'errors': errors
            }
            
        except Exception as e:
            logger.error(f"Error validating employee data: {e}")
            return {
                'is_valid': False,
                'errors': {'general': f"خطأ في التحقق من البيانات: {e}"}
            }
    
    def get_employee_hierarchy(self, employee_id):
        """الحصول على الهيكل الهرمي للموظف"""
        try:
            from ..models import Employee
            
            employee = Employee.objects.select_related(
                'direct_manager'
            ).prefetch_related('subordinates').get(id=employee_id)
            
            # Get hierarchy upwards (managers)
            hierarchy_up = []
            current = employee.direct_manager
            while current:
                hierarchy_up.append({
                    'id': current.id,
                    'name': current.full_name,
                    'position': current.job_position.title,
                    'level': len(hierarchy_up) + 1
                })
                current = current.direct_manager
            
            # Get hierarchy downwards (subordinates)
            def get_subordinates_recursive(emp, level=1):
                subordinates = []
                for sub in emp.subordinates.filter(is_active=True):
                    sub_data = {
                        'id': sub.id,
                        'name': sub.full_name,
                        'position': sub.job_position.title,
                        'level': level,
                        'subordinates': get_subordinates_recursive(sub, level + 1)
                    }
                    subordinates.append(sub_data)
                return subordinates
            
            hierarchy_down = get_subordinates_recursive(employee)
            
            return {
                'employee': {
                    'id': employee.id,
                    'name': employee.full_name,
                    'position': employee.job_position.title
                },
                'managers': list(reversed(hierarchy_up)),  # Top to bottom
                'subordinates': hierarchy_down,
                'total_subordinates': self._count_all_subordinates(employee)
            }
            
        except Employee.DoesNotExist:
            raise ValidationError("الموظف غير موجود")
        except Exception as e:
            logger.error(f"Error getting employee hierarchy: {e}")
            raise ValidationError(f"خطأ في جلب الهيكل الهرمي: {e}")
    
    def _count_all_subordinates(self, employee):
        """عد جميع المرؤوسين بشكل تكراري"""
        count = 0
        for sub in employee.subordinates.filter(is_active=True):
            count += 1 + self._count_all_subordinates(sub)
        return count
    
    def calculate_employee_cost(self, employee_id, include_benefits=True):
        """حساب التكلفة الإجمالية للموظف"""
        try:
            from ..models import Employee
            
            employee = Employee.objects.get(id=employee_id)
            
            # Basic salary (monthly)
            basic_salary = employee.basic_salary
            
            # Calculate annual basic salary
            annual_basic = basic_salary * 12
            
            cost_breakdown = {
                'basic_salary_monthly': float(basic_salary),
                'basic_salary_annual': float(annual_basic),
                'benefits': {},
                'total_monthly': float(basic_salary),
                'total_annual': float(annual_basic)
            }
            
            if include_benefits:
                # Add common benefits (these would come from benefit models)
                benefits = {
                    'social_insurance': float(basic_salary * Decimal('0.18')),  # 18% employer contribution
                    'medical_insurance': float(basic_salary * Decimal('0.05')),  # 5% medical
                    'transportation': 500.0,  # Fixed transportation allowance
                    'housing': float(basic_salary * Decimal('0.25')) if employee.housing_allowance else 0,
                }
                
                cost_breakdown['benefits'] = benefits
                
                # Calculate totals with benefits
                monthly_benefits = sum(benefits.values())
                cost_breakdown['total_monthly'] = float(basic_salary) + monthly_benefits
                cost_breakdown['total_annual'] = cost_breakdown['total_monthly'] * 12
            
            return cost_breakdown
            
        except Employee.DoesNotExist:
            raise ValidationError("الموظف غير موجود")
        except Exception as e:
            logger.error(f"Error calculating employee cost: {e}")
            raise ValidationError(f"خطأ في حساب تكلفة الموظف: {e}")
    
    def get_employees_expiring_documents(self, days_ahead=30):
        """الحصول على الموظفين الذين تنتهي وثائقهم قريباً"""
        try:
            from ..models import Employee
            
            expiry_date = date.today() + timedelta(days=days_ahead)
            
            # This would need to be implemented based on your document models
            employees_with_expiring_docs = []
            
            # Example implementation (you'd need to adjust based on your models)
            employees = Employee.objects.filter(is_active=True)
            
            for employee in employees:
                expiring_docs = []
                
                # Check passport expiry
                if employee.passport_expiry_date and employee.passport_expiry_date <= expiry_date:
                    expiring_docs.append({
                        'type': 'passport',
                        'expiry_date': employee.passport_expiry_date,
                        'days_remaining': (employee.passport_expiry_date - date.today()).days
                    })
                
                # Check ID expiry
                if employee.national_id_expiry and employee.national_id_expiry <= expiry_date:
                    expiring_docs.append({
                        'type': 'national_id',
                        'expiry_date': employee.national_id_expiry,
                        'days_remaining': (employee.national_id_expiry - date.today()).days
                    })
                
                # Check work permit expiry
                if employee.work_permit_expiry and employee.work_permit_expiry <= expiry_date:
                    expiring_docs.append({
                        'type': 'work_permit',
                        'expiry_date': employee.work_permit_expiry,
                        'days_remaining': (employee.work_permit_expiry - date.today()).days
                    })
                
                if expiring_docs:
                    employees_with_expiring_docs.append({
                        'employee': employee,
                        'expiring_documents': expiring_docs
                    })
            
            return employees_with_expiring_docs
            
        except Exception as e:
            logger.error(f"Error getting employees with expiring documents: {e}")
            raise ValidationError(f"خطأ في جلب الموظفين ذوي الوثائق المنتهية الصلاحية: {e}")
    
    # =============================================================================
    # REPORT GENERATION METHODS
    # =============================================================================
    
    def _generate_detailed_list_report(self, employees, format):
        """إنتاج تقرير قائمة مفصلة"""
        if format == 'pdf':
            return self._generate_pdf_report(employees, 'detailed_list')
        elif format == 'excel':
            return self._export_to_excel(employees)
        else:
            return self._export_to_csv(employees)
    
    def _generate_summary_report(self, employees, format):
        """إنتاج تقرير ملخص"""
        # Implementation for summary report
        pass
    
    def _generate_birthday_report(self, employees, format):
        """إنتاج تقرير أعياد الميلاد"""
        # Filter employees with birthdays in current month
        current_month = date.today().month
        birthday_employees = employees.filter(birth_date__month=current_month)
        
        if format == 'excel':
            return self._export_birthday_excel(birthday_employees)
        else:
            return self._export_birthday_csv(birthday_employees)
    
    def _generate_service_years_report(self, employees, format):
        """إنتاج تقرير سنوات الخدمة"""
        # Implementation for service years report
        pass
    
    def _generate_pdf_report(self, employees, report_type):
        """إنتاج تقرير PDF"""
        # This would require a PDF library like ReportLab
        # Implementation would go here
        pass
    
    def _export_birthday_excel(self, employees):
        """تصدير تقرير أعياد الميلاد إلى Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "أعياد الميلاد"
        
        # Headers
        headers = [
            'رقم الموظف', 'الاسم الكامل', 'القسم', 'المنصب',
            'تاريخ الميلاد', 'العمر', 'يوم الميلاد'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        for row, employee in enumerate(employees, 2):
            age = (date.today() - employee.birth_date).days // 365 if employee.birth_date else 0
            birthday_this_year = employee.birth_date.replace(year=date.today().year) if employee.birth_date else None
            
            ws.cell(row=row, column=1, value=employee.employee_number)
            ws.cell(row=row, column=2, value=employee.full_name)
            ws.cell(row=row, column=3, value=employee.department.name)
            ws.cell(row=row, column=4, value=employee.job_position.title)
            ws.cell(row=row, column=5, value=employee.birth_date)
            ws.cell(row=row, column=6, value=age)
            ws.cell(row=row, column=7, value=birthday_this_year)
        
        return wb
    
    def _export_birthday_csv(self, employees):
        """تصدير تقرير أعياد الميلاد إلى CSV"""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            'رقم الموظف', 'الاسم الكامل', 'القسم', 'المنصب',
            'تاريخ الميلاد', 'العمر', 'يوم الميلاد'
        ])
        
        # Data rows
        for employee in employees:
            age = (date.today() - employee.birth_date).days // 365 if employee.birth_date else 0
            birthday_this_year = employee.birth_date.replace(year=date.today().year) if employee.birth_date else None
            
            writer.writerow([
                employee.employee_number,
                employee.full_name,
                employee.department.name,
                employee.job_position.title,
                employee.birth_date,
                age,
                birthday_this_year
            ])
        
        return output.getvalue()