# =============================================================================
# ElDawliya HR Management System - New Employee Views
# =============================================================================
# Views for employee management (Employee, Documents, Emergency Contacts)
# Supports RTL Arabic interface and modern Django patterns
# =============================================================================

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.db.models import Q, Count, Sum, Avg
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.db import transaction
from django.utils import timezone
from datetime import date, timedelta
import json

from Hr.models import (
    Employee, Company, Branch, Department, JobPosition,
    EmployeeDocument, EmployeeEmergencyContact, EmployeeTraining, EmployeeNote
)
from Hr.forms.new_employee_forms import (
    NewEmployeeForm, EmployeeDocumentForm, EmployeeEmergencyContactForm
)


# =============================================================================
# EMPLOYEE VIEWS
# =============================================================================

class NewEmployeeListView(LoginRequiredMixin, ListView):
    """عرض قائمة الموظفين الجديد"""
    model = Employee
    template_name = 'Hr/new_employee/employee_list.html'
    context_object_name = 'employees'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = Employee.objects.select_related(
            'company', 'branch', 'department', 'job_position', 'direct_manager'
        ).order_by('employee_number')
        
        # التصفية حسب الشركة
        company_id = self.request.GET.get('company')
        if company_id:
            queryset = queryset.filter(company_id=company_id)
        
        # التصفية حسب الفرع
        branch_id = self.request.GET.get('branch')
        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)
        
        # التصفية حسب القسم
        department_id = self.request.GET.get('department')
        if department_id:
            queryset = queryset.filter(department_id=department_id)
        
        # التصفية حسب الحالة
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # التصفية حسب نوع التوظيف
        employment_type = self.request.GET.get('employment_type')
        if employment_type:
            queryset = queryset.filter(employment_type=employment_type)
        
        # البحث
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(employee_number__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(full_name__icontains=search) |
                Q(email__icontains=search) |
                Q(national_id__icontains=search) |
                Q(phone_primary__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إدارة الموظفين')
        
        # خيارات التصفية
        context['companies'] = Company.objects.filter(is_active=True)
        context['branches'] = Branch.objects.filter(is_active=True)
        context['departments'] = Department.objects.filter(is_active=True)
        context['status_choices'] = Employee.STATUS_CHOICES
        context['employment_type_choices'] = Employee.EMPLOYMENT_TYPE_CHOICES
        
        # القيم المحددة في التصفية
        context['search_value'] = self.request.GET.get('search', '')
        context['company_filter'] = self.request.GET.get('company', '')
        context['branch_filter'] = self.request.GET.get('branch', '')
        context['department_filter'] = self.request.GET.get('department', '')
        context['status_filter'] = self.request.GET.get('status', '')
        context['employment_type_filter'] = self.request.GET.get('employment_type', '')
        
        # إحصائيات
        all_employees = Employee.objects.all()
        context['total_employees'] = all_employees.count()
        context['active_employees'] = all_employees.filter(status='active').count()
        context['inactive_employees'] = all_employees.filter(status='inactive').count()
        context['on_leave_employees'] = all_employees.filter(status='on_leave').count()
        
        return context


class NewEmployeeDetailView(LoginRequiredMixin, DetailView):
    """عرض تفاصيل الموظف الجديد"""
    model = Employee
    template_name = 'Hr/new_employee/employee_detail.html'
    context_object_name = 'employee'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee = self.get_object()
        
        context['title'] = f"ملف الموظف - {employee.full_name}"
        
        # المعلومات الإضافية
        context['age'] = employee.age
        context['years_of_service'] = employee.years_of_service
        
        # الوثائق والمرفقات
        context['documents'] = employee.documents.all()[:5]
        context['emergency_contacts'] = employee.emergency_contacts.all()
        context['trainings'] = employee.trainings.all()[:5]
        context['notes'] = employee.notes.all()[:5]
        
        # إحصائيات الحضور (آخر 30 يوم)
        thirty_days_ago = date.today() - timedelta(days=30)
        context['recent_attendance'] = employee.attendance_records.filter(
            date__gte=thirty_days_ago
        ).count()
        
        # الإجازات الحالية
        context['current_leave_requests'] = employee.leave_requests.filter(
            status__in=['submitted', 'approved'],
            start_date__lte=date.today(),
            end_date__gte=date.today()
        )
        
        # آخر تقييم أداء
        context['latest_evaluation'] = employee.performance_evaluations.filter(
            status='completed'
        ).order_by('-evaluation_date').first()
        
        return context


class NewEmployeeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """إنشاء موظف جديد"""
    model = Employee
    form_class = NewEmployeeForm
    template_name = 'Hr/new_employee/employee_form.html'
    permission_required = 'Hr.add_employee'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إضافة موظف جديد')
        context['action'] = 'create'
        return context
    
    def form_valid(self, form):
        # تعيين المستخدم الذي أنشأ الموظف
        form.instance.created_by = self.request.user
        
        # إنشاء رقم موظف تلقائي إذا لم يتم تحديده
        if not form.instance.employee_number:
            company = form.instance.company
            last_employee = Employee.objects.filter(company=company).order_by('-employee_number').first()
            if last_employee and last_employee.employee_number.isdigit():
                next_number = int(last_employee.employee_number) + 1
            else:
                next_number = 1
            form.instance.employee_number = f"{company.code}{next_number:04d}"
        
        messages.success(self.request, f"تم إنشاء ملف الموظف '{form.instance.full_name}' بنجاح")
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('hr:new_employee_detail', kwargs={'pk': self.object.pk})


class NewEmployeeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """تحديث بيانات الموظف"""
    model = Employee
    form_class = NewEmployeeForm
    template_name = 'Hr/new_employee/employee_form.html'
    permission_required = 'Hr.change_employee'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"تحديث بيانات الموظف - {self.get_object().full_name}"
        context['action'] = 'update'
        return context
    
    def get_success_url(self):
        return reverse('hr:new_employee_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, f"تم تحديث بيانات الموظف '{form.instance.full_name}' بنجاح")
        return super().form_valid(form)


# =============================================================================
# EMPLOYEE DOCUMENT VIEWS
# =============================================================================

class NewEmployeeDocumentListView(LoginRequiredMixin, ListView):
    """عرض وثائق الموظف"""
    model = EmployeeDocument
    template_name = 'Hr/new_employee/employee_documents.html'
    context_object_name = 'documents'
    paginate_by = 20
    
    def get_queryset(self):
        employee_id = self.kwargs.get('employee_id')
        return EmployeeDocument.objects.filter(
            employee_id=employee_id
        ).order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee_id = self.kwargs.get('employee_id')
        context['employee'] = get_object_or_404(Employee, pk=employee_id)
        context['title'] = f"وثائق الموظف - {context['employee'].full_name}"
        return context


class NewEmployeeDocumentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """رفع وثيقة جديدة للموظف"""
    model = EmployeeDocument
    form_class = EmployeeDocumentForm
    template_name = 'Hr/new_employee/employee_document_form.html'
    permission_required = 'Hr.add_employeedocument'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        employee_id = self.kwargs.get('employee_id')
        kwargs['employee'] = get_object_or_404(Employee, pk=employee_id)
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee_id = self.kwargs.get('employee_id')
        context['employee'] = get_object_or_404(Employee, pk=employee_id)
        context['title'] = f"رفع وثيقة جديدة - {context['employee'].full_name}"
        return context
    
    def form_valid(self, form):
        form.instance.uploaded_by = self.request.user
        messages.success(self.request, _('تم رفع الوثيقة بنجاح'))
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('hr:new_employee_documents', kwargs={'employee_id': self.kwargs.get('employee_id')})


# =============================================================================
# EMPLOYEE EMERGENCY CONTACT VIEWS
# =============================================================================

class NewEmployeeEmergencyContactCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """إضافة جهة اتصال طوارئ للموظف"""
    model = EmployeeEmergencyContact
    form_class = EmployeeEmergencyContactForm
    template_name = 'Hr/new_employee/emergency_contact_form.html'
    permission_required = 'Hr.add_employeeemergencycontact'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        employee_id = self.kwargs.get('employee_id')
        kwargs['employee'] = get_object_or_404(Employee, pk=employee_id)
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee_id = self.kwargs.get('employee_id')
        context['employee'] = get_object_or_404(Employee, pk=employee_id)
        context['title'] = f"إضافة جهة اتصال طوارئ - {context['employee'].full_name}"
        return context
    
    def form_valid(self, form):
        # إذا كانت جهة الاتصال الأساسية، قم بإلغاء الأساسية من الأخريات
        if form.instance.is_primary:
            EmployeeEmergencyContact.objects.filter(
                employee=form.instance.employee,
                is_primary=True
            ).update(is_primary=False)
        
        messages.success(self.request, _('تم إضافة جهة الاتصال بنجاح'))
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('hr:new_employee_detail', kwargs={'pk': self.kwargs.get('employee_id')})


# =============================================================================
# AJAX VIEWS
# =============================================================================

@login_required
def new_employee_search_ajax(request):
    """البحث عن الموظفين - AJAX"""
    query = request.GET.get('q', '')
    employees = []
    
    if len(query) >= 2:
        employees_qs = Employee.objects.filter(
            Q(employee_number__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(full_name__icontains=query) |
            Q(email__icontains=query)
        ).filter(status='active')[:10]
        
        employees = [
            {
                'id': emp.id,
                'employee_number': emp.employee_number,
                'full_name': emp.full_name,
                'department': emp.department.name if emp.department else '',
                'job_position': emp.job_position.title if emp.job_position else '',
                'photo_url': emp.photo.url if emp.photo else None
            }
            for emp in employees_qs
        ]
    
    return JsonResponse({'employees': employees})


@login_required
def new_employee_quick_stats(request, pk):
    """إحصائيات سريعة للموظف - AJAX"""
    employee = get_object_or_404(Employee, pk=pk)
    
    # حساب الإحصائيات
    current_month = date.today().replace(day=1)
    next_month = (current_month + timedelta(days=32)).replace(day=1)
    
    stats = {
        'age': employee.age,
        'years_of_service': employee.years_of_service,
        'documents_count': employee.documents.count(),
        'emergency_contacts_count': employee.emergency_contacts.count(),
        'current_month_attendance': employee.attendance_records.filter(
            date__gte=current_month,
            date__lt=next_month
        ).count(),
        'pending_leave_requests': employee.leave_requests.filter(
            status='submitted'
        ).count(),
        'approved_leave_requests': employee.leave_requests.filter(
            status='approved',
            start_date__gte=date.today()
        ).count()
    }
    
    return JsonResponse(stats)


@login_required
def new_employee_status_update(request, pk):
    """تحديث حالة الموظف - AJAX"""
    if request.method == 'POST':
        employee = get_object_or_404(Employee, pk=pk)
        new_status = request.POST.get('status')
        
        if new_status in dict(Employee.STATUS_CHOICES):
            old_status = employee.get_status_display()
            employee.status = new_status
            employee.save()
            
            # إضافة ملاحظة عن تغيير الحالة
            EmployeeNote.objects.create(
                employee=employee,
                note_type='general',
                title=f'تغيير حالة الموظف',
                content=f'تم تغيير حالة الموظف من "{old_status}" إلى "{employee.get_status_display()}"',
                created_by=request.user
            )
            
            messages.success(request, f'تم تحديث حالة الموظف إلى "{employee.get_status_display()}"')
            return JsonResponse({'success': True, 'new_status': employee.get_status_display()})
    
    return JsonResponse({'success': False, 'error': 'طلب غير صحيح'})


# =============================================================================
# BULK OPERATIONS
# =============================================================================

@login_required
def new_employee_bulk_update(request):
    """تحديث مجموعي للموظفين"""
    if request.method == 'POST':
        employee_ids = request.POST.getlist('employee_ids')
        action = request.POST.get('action')
        
        if not employee_ids:
            messages.error(request, _('يرجى اختيار موظف واحد على الأقل'))
            return redirect('hr:new_employee_list')
        
        employees = Employee.objects.filter(id__in=employee_ids)
        
        if action == 'activate':
            employees.update(status='active')
            messages.success(request, f'تم تفعيل {employees.count()} موظف')
        
        elif action == 'deactivate':
            employees.update(status='inactive')
            messages.success(request, f'تم إلغاء تفعيل {employees.count()} موظف')
        
        elif action == 'export':
            # تصدير بيانات الموظفين
            return new_export_employees_excel(request, employee_ids)
        
        return redirect('hr:new_employee_list')
    
    return redirect('hr:new_employee_list')


def new_export_employees_excel(request, employee_ids):
    """تصدير بيانات الموظفين إلى Excel"""
    import openpyxl
    from django.http import HttpResponse
    
    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "بيانات الموظفين"
    
    # إضافة العناوين
    headers = [
        'رقم الموظف', 'الاسم الكامل', 'البريد الإلكتروني', 'الهاتف',
        'الشركة', 'الفرع', 'القسم', 'المنصب', 'تاريخ التوظيف', 'الحالة'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # إضافة بيانات الموظفين
    employees = Employee.objects.filter(id__in=employee_ids).select_related(
        'company', 'branch', 'department', 'job_position'
    )
    
    for row, employee in enumerate(employees, 2):
        ws.cell(row=row, column=1, value=employee.employee_number)
        ws.cell(row=row, column=2, value=employee.full_name)
        ws.cell(row=row, column=3, value=employee.email)
        ws.cell(row=row, column=4, value=employee.phone_primary)
        ws.cell(row=row, column=5, value=employee.company.name)
        ws.cell(row=row, column=6, value=employee.branch.name if employee.branch else '')
        ws.cell(row=row, column=7, value=employee.department.name if employee.department else '')
        ws.cell(row=row, column=8, value=employee.job_position.title if employee.job_position else '')
        ws.cell(row=row, column=9, value=employee.hire_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=10, value=employee.get_status_display())
    
    # إعداد الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="employees.xlsx"'
    
    wb.save(response)
    return response
